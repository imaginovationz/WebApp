# backend/ROIExcelDownload.py
import os
import sqlite3
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

TEMPLATE_XLSX = "ROI_Portfolio_Project Intake number and name_version.xlsx"

def _db_conn(db_path: str) -> sqlite3.Connection:
    #conn = sqlite3.connect(db_path)
    conn = sqlite3.connect(db_path, timeout=30, check_same_thread=False)
    
    conn.row_factory = sqlite3.Row
    # Make SQLite more concurrent-friendly
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA busy_timeout=5000;")      # wait up to 5s instead of failing
    conn.execute("PRAGMA synchronous=NORMAL;")     # good default for WAL

    return conn

def _first_nonempty(iterable, default=""):
    for v in iterable:
        if v:
            return v
    return default

def _derive_roi_month(payload: dict) -> str:
    """
    Resolve the ROI month to use when generating the workbook:
    1) If UI sends it explicitly (roi_month), use that.
    2) Else derive from posted UI rows (auto_rows_by_sub / tdm_rows_by_sub → Dropdown2).
    3) Else fall back to current YYYY-MM.
    """
    # >>> CHANGE: accept explicit month from the UI payload
    direct = (payload.get("roi_month") or payload.get("ROIMonth") or payload.get("ROI_month") or "").strip()
    if direct:
        return direct.replace("/", "-")[:7]

    # >>> CHANGE: also look into the actual keys the UI posts and the 'Dropdown2' field
    containers = (
        "auto_rows_by_sub", "tdm_rows_by_sub",           # UI posts these maps
        "autoRowsByButton", "tdmRowsByButton",           # legacy names (kept)
        "automationRows", "tdmRows", "rows"              # generic fallbacks
    )

    def _norm_month(v):
        s = str(v or "").strip()
        if not s:
            return ""
        # accept YYYY-MM, YYYY/MM, YYYY-MM-DD, etc; normalize to 'YYYY-MM'
        return s.replace("/", "-")[:7]

    for key in containers:
        bag = payload.get(key)
        if isinstance(bag, dict):
            for _, rows in (bag or {}).items():
                for r in rows or []:
                    cand = _norm_month(
                        r.get("Dropdown2")
                        or r.get("ROI_month") or r.get("ROIMonth") or r.get("roi_month")
                    )
                    if cand:
                        return cand
        elif isinstance(bag, list):
            for r in bag or []:
                cand = _norm_month(
                    r.get("Dropdown2")
                    or r.get("ROI_month") or r.get("ROIMonth") or r.get("roi_month")
                )
                if cand:
                    return cand

    # Fallback: current month (YYYY-MM)
    return datetime.now().strftime("%Y-%m")

def build_roi_workbook(payload: dict, db_path: str):
    """
    Build an Excel workbook from the fixed template and populate only the 'Summary' sheet
    as per the latest instructions. Returns (BytesIO, download_filename).
    """
    intake_number = int(payload.get("intake_number") or payload.get("intakeNo") or 0)
    intake_name = (payload.get("intake_name") or payload.get("intakeName") or "").strip()

    # Simple UI fields (safe fallbacks)
    release_name = (payload.get("ReleaseName") or payload.get("release") or "").strip()
    applications_impacted = (payload.get("ApplicationsImpacted") or payload.get("applications_impacted") or "").strip()
    automation_lead = (payload.get("AutomationLeadName") or payload.get("automation_lead") or "").strip()
    functional_qe_lead = (payload.get("Functional_qe_lead") or payload.get("functional_qe_lead") or "").strip()

    roi_month = _derive_roi_month(payload)

    # Locate the template in the required folder(s)
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(here, "files", "templates", TEMPLATE_XLSX),
        os.path.join(here, "templates", TEMPLATE_XLSX),
        os.path.join(here, TEMPLATE_XLSX),
    ]
    template_path = next((p for p in candidates if os.path.exists(p)), None)
    if not template_path:
        raise FileNotFoundError(f"Template not found: {TEMPLATE_XLSX}")

    wb = load_workbook(template_path)
    ws = wb["Summary"]  # we only touch Summary per this step

    # Fill static details
    ws["B2"].value = f"{intake_number} - {intake_name}"     # Project Number and Name
    ws["B3"].value = release_name                           # Release
    ws["B4"].value = applications_impacted                  # Applications Impacted
    ws["B5"].value = automation_lead                        # Automation Lead
    ws["B6"].value = functional_qe_lead                     # Functional Lead
    ws["B9"].value = roi_month                              # ROI Reporting Month

    conn = _db_conn(db_path)
    cur = conn.cursor()

    cur.execute(
        """
        SELECT functional_qe_lead
          FROM projects
         WHERE intake_number = ? AND intake_name = ?
         ORDER BY COALESCE(timestamp, 0) DESC
         LIMIT 1
        """,
        (intake_number, intake_name)
    )
    proj_row = cur.fetchone()
    functional_qe_lead = (proj_row["functional_qe_lead"] if proj_row else "") or ""
    ws["B6"].value = functional_qe_lead

    # Distinct TranInitiative → B7, C7, D7...
    cur.execute("""
        SELECT DISTINCT TRIM(TranInitiative) AS ti
        FROM projectroi
        WHERE intake_number = ?
          AND intake_name   = ?
          AND COALESCE(ROI_month,'') = ?
          AND TRIM(COALESCE(TranInitiative,'')) <> ''
        ORDER BY ti
    """, (intake_number, intake_name, roi_month))
    tis = [r["ti"] for r in cur.fetchall() if r["ti"]]
    for idx, val in enumerate(tis):
        ws.cell(row=7, column=2 + idx, value=val)

    # Distinct "Other Utilities" from AutomationFrmk → B8, C8, D8...
    cur.execute("""
        SELECT DISTINCT TRIM(AutomationFrmk) AS af
        FROM projectroi
        WHERE intake_number = ?
          AND intake_name   = ?
          AND COALESCE(ROI_month,'') = ?
          AND AutomationFrmk LIKE ?
    """, (intake_number, intake_name, roi_month, "%Other Utilities%"))
    utils = []
    for r in cur.fetchall():
        af = r["af"] or ""
        # Expect values like "Other Utilities - <UtilityName>"
        if "-" in af:
            utils.append(af.split("-", 1)[1].strip())
        else:
            utils.append(af.strip())
    for idx, val in enumerate(sorted(set([u for u in utils if u]))):
        ws.cell(row=8, column=2 + idx, value=val)

    ws_nonsoa_template = wb["NonSOA"]
    ws_soa_template = wb["SOA"]


    # Aggregations for Summary (B12:C16)
    tdm_categories = (
        "mmtg Deal creation",
        "CLASS deal creation",
        "Savvy Deal creation",
        "Other Deal/Test data creation utility",
    )

    def sum_pair(extra_where, params_tail):
        sql = f"""
            SELECT COALESCE(SUM(ROI),0.0) AS sum_roi,
                   COALESCE(SUM(SavingsPD),0.0) AS sum_pd
            FROM projectroi
            WHERE intake_number=? AND intake_name=? AND COALESCE(ROI_month,'')=?
              AND {extra_where}
        """
        cur.execute(sql, (intake_number, intake_name, roi_month, *params_tail))
        row = cur.fetchone()
        return float(row["sum_roi"] or 0.0), float(row["sum_pd"] or 0.0)

    # B13/C13: TDM buckets; exclude "Other Utilities"
    placeholders = ",".join(["?"] * len(tdm_categories))
    b13, c13 = sum_pair(
        f"SavingsCategory IN ({placeholders}) AND (AutomationFrmk NOT LIKE ?)",
        [*tdm_categories, "%Other Utilities%"]
    )

    # B14/C14: TranInitiative blank; not TDM; exclude "Other Utilities"
    b14, c14 = sum_pair(
        f"(TRIM(COALESCE(TranInitiative,''))='') "
        f"AND (SavingsCategory NOT IN ({placeholders})) "
        f"AND (AutomationFrmk NOT LIKE ?)",
        [*tdm_categories, "%Other Utilities%"]
    )

    # B15/C15: TranInitiative present; not TDM
    b15, c15 = sum_pair(
        f"(TRIM(COALESCE(TranInitiative,''))<>'') "
        f"AND (SavingsCategory NOT IN ({placeholders}))",
        [*tdm_categories]
    )

    # B16/C16: "Other Utilities" and TranInitiative blank
    b16, c16 = sum_pair(
        f"(AutomationFrmk LIKE ?) AND (TRIM(COALESCE(TranInitiative,''))='')",
        ["%Other Utilities%"]
    )

    '''
    b17, c17 = sum_pair(
    #logic to be implemented for Adhoc utilities
            b17 = 0,
            c17 = 0
    )
    '''
    # Write the numbers
    ws["B13"].value = round(b13, 2); ws["C13"].value = round(c13, 2)
    ws["B14"].value = round(b14, 2); ws["C14"].value = round(c14, 2)
    ws["B15"].value = round(b15, 2); ws["C15"].value = round(c15, 2)
    ws["B16"].value = round(b16, 2); ws["C16"].value = round(c16, 2)
    #ws["B17"].value = round(b17, 2); ws["C17"].value = round(c17, 2)
    ws["B12"].value = round(b13 + b14 + b15 + b16 , 2)
    ws["C12"].value = round(c13 + c14 + c15 + c16 , 2)

    cur.execute(
        f"""
        SELECT
            TRIM(COALESCE(SavingsCategory,'')) AS sc,
            TRIM(COALESCE(AutomationFrmk,''))  AS af,
            COALESCE(PD1_TotalManualPDsNByN1,0.0) * COALESCE(N1_AvgManualPD,0.0) AS pd_prod
        FROM projectroi
        WHERE intake_number=? AND intake_name=? AND COALESCE(ROI_month,'')=?
        AND SavingsCategory IN ({placeholders})
        ORDER BY ROWID
        """,
        # ##: dropped the '%Other Utilities%' parameter; only pass TDM categories now
        (intake_number, intake_name, roi_month, *tdm_categories)
    )
    l_rows = cur.fetchall() or []
    start_row = 21

    x = len(l_rows)  # number of rows to be populated for TDM
    if x > 1:
       ws.insert_rows(idx=start_row + 1, amount=x - 1)  # insert at row 22

    for idx, r in enumerate(l_rows):

        label = r["sc"]
        if r["af"]:
            label = f"{label} ({r['af']})"
        ws.cell(row=start_row + idx, column=2, value=label)  # Col B
        ws.cell(row=start_row + idx, column=3,               # Col C (PD product)
                value=float(round(r["pd_prod"] or 0.0, 2)))
        
        #ws.cell(row=start_row + idx, column=2, value=r["sc"])                 # Col B
        #ws.cell(row=start_row + idx, column=3, value=float(round(r["pd_prod"] or 0.0, 2)))  # Col C
    # ------------------------------------------------------------------------
    
    #print("the number of rows for test data is :", idx)
    
    
    # --- CHANGE [m1..m8]: Framework rows BY and category PD products --------
    #x = len(l_rows)  # number of rows printed for l1
    #y_start = 21 + x + 7  # "B21 + x + 3"

    y_start_base = 25                              # in template, Automation data starts at row 27
    y_start = y_start_base + max(x - 1, 0) 

    # Fetch DISTINCT frameworks as BY list with required filters
    cur.execute(
        f"""
        SELECT DISTINCT TRIM(COALESCE(AutomationFrmk,'')) AS af
        FROM projectroi
        WHERE intake_number=? AND intake_name=? AND COALESCE(ROI_month,'')=?
          AND TRIM(COALESCE(TranInitiative,''))=''                  -- initiative blank
          AND (AutomationFrmk NOT LIKE ?)                           -- exclude 'Other Utilities'
          AND (SavingsCategory NOT IN ({placeholders}))             -- exclude TDM buckets
        ORDER BY af
        """,
        (intake_number, intake_name, roi_month, "%Other Utilities%", *tdm_categories)
    )
    by_frameworks = [r["af"] for r in cur.fetchall() if r["af"]]

    # FIX: Insert blank rows for Automation section so we don't clobber the next section.
    n = len(by_frameworks)
    if n > 1:
       ws.insert_rows(idx=y_start + 1, amount=n - 1)  # make space below first data row

    # helper to compute PD product for a given framework + category name
    def sum_pd_for_framework_and_category(af_name: str, cat_exact: str,app_name: str = None) -> float:
        sql = """
            SELECT COALESCE(SUM(COALESCE(PD1_TotalManualPDsNByN1,0.0) * COALESCE(N1_AvgManualPD,0.0)),0.0) AS val
              FROM projectroi
             WHERE intake_number=? AND intake_name=? AND COALESCE(ROI_month,'')=?
               AND TRIM(COALESCE(TranInitiative,''))=''           -- initiative blank
               AND TRIM(COALESCE(AutomationFrmk,'')) = ?          -- this framework
               AND SavingsCategory = ?
            """
           # (intake_number, intake_name, roi_month, af_name, cat_exact)
        
        params = [intake_number, intake_name, roi_month, af_name, cat_exact]
        if app_name:
            sql += " AND TRIM(COALESCE(Application,'')) = ?"
            params.append(app_name)
        cur.execute(sql, params)
        row = cur.fetchone()
        return float(row["val"] or 0.0)

    CAT_NEW  = "New Test case Created"
    CAT_Update  = "Existing Test case Updated/Reused"
    CAT_DIT  = "Test Case Execution - DIT"
    CAT_SIT  = "Test Case Execution - SIT"
    CAT_UAT  = "Test Case Execution - UAT"
    CAT_RB   = "Test Case Execution - RollBack"
    CAT_RF   = "Test Case Execution - RollForward"
    CAT_ADH  = "Test Case Execution - Adhoc/utility"

    def pick_application_for_framework(af_name: str) -> str:
        """
        Return the most frequent non-empty Application for this framework within the BY filters:
          - ROI month matches
          - initiative blank
          - NOT 'Other Utilities'
          - NOT in TDM buckets
        """
        cur.execute(
            f"""
            SELECT TRIM(COALESCE(Application,'')) AS app, COUNT(*) AS cnt
              FROM projectroi
             WHERE intake_number=? AND intake_name=? AND COALESCE(ROI_month,'')=?
               AND TRIM(COALESCE(TranInitiative,''))=''
               AND (AutomationFrmk NOT LIKE ?)
               AND (SavingsCategory NOT IN ({placeholders}))
               AND TRIM(COALESCE(Application,'')) <> ''
               AND TRIM(COALESCE(AutomationFrmk,'')) = ?
             GROUP BY app
             ORDER BY cnt DESC, app ASC
             LIMIT 1
            """,
            (intake_number, intake_name, roi_month, "%Other Utilities%", *tdm_categories, af_name)
        )
        rr = cur.fetchone()
        return (rr["app"] if rr else "") or ""
    
    for i, af in enumerate(by_frameworks):
        r = y_start + i
        # Column B(Y+i): framework name
        ws.cell(row=r, column=2, value=af)
        
        # Application
        app = pick_application_for_framework(af)
        ws.cell(row=r, column=3, value=app)

        # Column D(Y+i): New Test case Created
        ws.cell(row=r, column=4, value=round(sum_pd_for_framework_and_category(af, CAT_NEW), 2))
        
        # Column E(Y+i): Existing Test case Updated/Reused
        ws.cell(row=r, column=5, value=round(sum_pd_for_framework_and_category(af, CAT_Update), 2))

        # Column F(Y+i): DIT
        ws.cell(row=r, column=6, value=round(sum_pd_for_framework_and_category(af, CAT_DIT), 2))
        
        # Column G(Y+i): SIT
        ws.cell(row=r, column=7, value=round(sum_pd_for_framework_and_category(af, CAT_SIT), 2))
        
        # Column H(Y+i): UAT
        ws.cell(row=r, column=8, value=round(sum_pd_for_framework_and_category(af, CAT_UAT), 2))
        
        # Column I(Y+i): RollBack
        ws.cell(row=r, column=9, value=round(sum_pd_for_framework_and_category(af, CAT_RB), 2))
        
        # Column J(Y+i): RollForward
        ws.cell(row=r, column=10, value=round(sum_pd_for_framework_and_category(af, CAT_RF), 2))
        
        # Column K(Y+i): Adhoc/utility
        ws.cell(row=r, column=11, value=round(sum_pd_for_framework_and_category(af, CAT_ADH), 2))

    #print("the number of rows for automation rows is :", i)
    #last_idx_l = (len(l_rows) - 1) if len(l_rows) > 0 else 0
    #last_idx_bf = (len(by_frameworks) - 1) if len(by_frameworks) > 0 else 0
    #z = 21 + 12 + last_idx_l + last_idx_bf  # row Z

    last_idx_l = (x - 1) if x > 0 else 0
    last_idx_bf = (n - 1) if n > 0 else 0
    z = 31 + last_idx_l + last_idx_bf  # first data row for Transformation after shifts 38
    
    cur.execute(
        f"""
        SELECT
            TRIM(COALESCE(TranInitiative,''))  AS ti,
            TRIM(COALESCE(AutomationFrmk,''))  AS af,
            TRIM(COALESCE(Application,''))     AS app
        FROM projectroi
        WHERE intake_number=? AND intake_name=? AND COALESCE(ROI_month,'')=?
        AND (SavingsCategory NOT IN ({placeholders}))
        AND (AutomationFrmk NOT LIKE ?)
        AND TRIM(COALESCE(TranInitiative,''))<>''
        GROUP BY ti, af, app
        ORDER BY ti, af, app
        """,
        (intake_number, intake_name, roi_month, *tdm_categories, "%Other Utilities%")
    )
    z_rows = cur.fetchall() or []

    def _sum_pd_for_ti_af_cat(ti: str, af: str, cat: str) -> float:
        cur.execute(
            """
            SELECT COALESCE(SUM(COALESCE(PD1_TotalManualPDsNByN1,0.0) * COALESCE(N1_AvgManualPD,0.0)),0.0) AS val
            FROM projectroi
            WHERE intake_number=? AND intake_name=? AND COALESCE(ROI_month,'')=?
            AND TRIM(COALESCE(TranInitiative,'')) = ?
            AND TRIM(COALESCE(AutomationFrmk,'')) = ?
            AND UPPER(SavingsCategory) = UPPER(?)
            """,
            (intake_number, intake_name, roi_month, ti, af, cat)
        )
        row = cur.fetchone()
        return float(row["val"] or 0.0)

    CAT_NEW   = "New Test case Created"
    CAT_UPD   = "Existing Test case Updated/Reused"
    CAT_DIT   = "TEST CASE EXECUTION - DIT"
    CAT_SIT   = "TEST CASE EXECUTION - SIT"
    CAT_UAT   = "TEST CASE EXECUTION - UAT"
    CAT_RB    = "Test Case Execution - RollBack"
    CAT_RF   = "Test Case Execution - RollForward"     # per spec: JZ uses RollBack again
    CAT_ADH   = "Test Case Execution - Adhoc/utility"

    for off, r in enumerate(z_rows):
        rr = z + off
        ti = r["ti"]; af = r["af"]; app = r["app"]

        # 2a: BZ = "TranInitiative (AutomationFrmk)"
        ws.cell(row=rr, column=2, value=f"{ti} ({af})")

        # 2a.1: CZ = Application
        ws.cell(row=rr, column=3, value=app)

        # 2a.2..2a.8: DZ..KZ = PD1_TotalManualPDsNByN1 * N1_AvgManualPD per category
        ws.cell(row=rr, column=4,  value=round(_sum_pd_for_ti_af_cat(ti, af, CAT_NEW), 2))  # DZ
        ws.cell(row=rr, column=5,  value=round(_sum_pd_for_ti_af_cat(ti, af, CAT_UPD), 2))  # EZ
        ws.cell(row=rr, column=6,  value=round(_sum_pd_for_ti_af_cat(ti, af, CAT_DIT), 2))  # FZ
        ws.cell(row=rr, column=7,  value=round(_sum_pd_for_ti_af_cat(ti, af, CAT_SIT), 2))  # GZ
        ws.cell(row=rr, column=8,  value=round(_sum_pd_for_ti_af_cat(ti, af, CAT_UAT), 2))  # HZ
        ws.cell(row=rr, column=9,  value=round(_sum_pd_for_ti_af_cat(ti, af, CAT_RB),  2))  # IZ
        ws.cell(row=rr, column=10, value=round(_sum_pd_for_ti_af_cat(ti, af, CAT_RF), 2))  # JZ (RollBack again as specified)
        ws.cell(row=rr, column=11, value=round(_sum_pd_for_ti_af_cat(ti, af, CAT_ADH), 2))  # KZ
        # === END NEW (Z-block) ======================================================

    #print("automation start row y_start ", y_start) = 34
    #print("automation first row r ", r) == object
    #print("z is ", z) = 38
    #print("the starting row for transformation rr is :", rr) = 39
    

    try:
        tdm_cats = set(tdm_categories)
    except Exception:
        tdm_cats = {
            "mmtg Deal creation",
            "CLASS deal creation",
            "Savvy Deal creation",
            "Other Deal/Test data creation utility",
        }

    # Get template sheets if present (we DO NOT remove them later; zero impact to your current flow)
    try:
        ws_nonsoa_template = wb["NonSOA"]
    except KeyError:
        ws_nonsoa_template = None

    try:
        ws_soa_template = wb["SOA"]
    except KeyError:
        ws_soa_template = None

    def _safe_title(raw: str) -> str:

        bad = set(':\\/?*[]')
        t = (raw or "Sheet")
        t = "".join((" " if ch in bad else ch) for ch in t)
        t = " ".join(t.split()).strip() or "Sheet"
        # base trimmed to 31
        base = t[:31]

        # ensure unique within this workbook
        name = base
        i = 1
        while name in wb.sheetnames:
            suffix = f" ({i})"
            # keep total length ≤ 31
            name = (base[: 31 - len(suffix)]).rstrip() + suffix
            i += 1
        return name
    
    def _copy_template(tpl, title):
        """Copy an existing template sheet with a safe title."""
        if tpl is None:
            return None
        ws_new = wb.copy_worksheet(tpl)
        ws_new.title = _safe_title(title)
        return ws_new

    # Fetch all rows for this project & month once (only the columns we need)
    cur2 = conn.cursor()
    cur2.execute("""
        SELECT
            TRIM(COALESCE(AutomationFrmk,''))                 AS AutomationFrmk,
            TRIM(COALESCE(Application,''))                    AS Application,
            TRIM(COALESCE(SavingsCategory, [SavingsCategory], '')) AS SavingsCategory,

            (COALESCE(N1_AvgManualPD, 0.0) * COALESCE(PD1_TotalManualPDsNByN1, 0.0)) AS N,
            COALESCE(N1_AvgManualPD, 0.0)                  AS N1,
            COALESCE(PD1_TotalManualPDsNByN1, 0.0)        AS PD1,
            COALESCE(C1_TotalManualCost_D, 0.0) AS C1,
            COALESCE(N2_AvgAutomationPD, 0.0)              AS N2,
            COALESCE(PD2_TotalAutomatedPDsNByN2, 0.0)     AS PD2,
            COALESCE(C2_TotalAutomationCost_D, 0.0)          AS C2,
            COALESCE(N3_NumberOfCycles, 0.0)               AS N3,
            COALESCE(ROI, 0.0)                                 AS ROI,
            COALESCE(SavingsPD, 0.0)                           AS SavingsPD
        FROM projectroi
        
        WHERE intake_number = ? AND intake_name = ? AND substr(COALESCE(ROI_month,''),1,7) = ?
                 ORDER BY ROWID
    """, (intake_number, intake_name, roi_month))
    _all_rows_btn = [dict(r) for r in cur2.fetchall()]

    # Group rows by UI "button" (AutomationFrmk)
    _by_button = {}
    for r in _all_rows_btn:
        key = (r["AutomationFrmk"] or "").strip()
        if not key:
            key = "UNSPECIFIED"
        _by_button.setdefault(key, []).append(r)


    def _write_nonsoa_sheet(button_name: str, rows: list):
        """
        Clone NonSOA template and fill two sections without altering the template layout:
        - TDM Savings from row 6 (A..L), contiguous
        - Automation Savings — Breakdown from row y = 6 + 9 + <TDM row count> (A..L), contiguous
        Totals:
        - AE9 = sum(Column K) of all TDM rows on this sheet
        -  E9 = sum(Column K) of all Automation rows on this sheet
        -  D9 = AE9 + E9
        """
        ws_btn = _copy_template(ws_nonsoa_template, button_name or "NonSOA")
        if ws_btn is None:
            return  # template missing; skip gracefully

        # Partition rows once; do NOT insert rows (preserve template merges/formatting)
        tdm_rows  = [r for r in rows if (r.get("SavingsCategory") in tdm_cats)]
        auto_rows = [r for r in rows if (r.get("SavingsCategory") not in tdm_cats)]

        tdm_count  = int(len(tdm_rows))      # <- do NOT assign the list itself here
        auto_count = int(len(auto_rows))     # <- same for auto

        

        if tdm_count > 0:
            ws_btn.insert_rows(idx=6, amount=tdm_count)

        
        
        tdm_band_start = 5
        tdm_band_end   = 5 + max(tdm_count - 1, 0) 
        auto_band_start = 11 + tdm_count
        auto_count      = int(len(auto_rows))
        auto_band_end   = auto_band_start + max(auto_count - 1, 0)

        #print("tdm rows :", tdm_count)
        #print("auto_band_start :",auto_band_start)

        # Unmerge only ranges that intersect our data bands AND columns A..L (1..12).
        # Skip invalid/single-cell merges; ignore any unmerge errors gracefully.
        for mcr in list(ws_btn.merged_cells.ranges):
            r1, r2 = mcr.min_row, mcr.max_row
            c1, c2 = mcr.min_col, mcr.max_col

            if r1 == r2 == 16:
                continue

            intersects_cols = not (c2 < 1 or c1 > 12)  # limit to data columns A..L
            in_tdm  = (tdm_count > 0) and (r2 >= tdm_band_start and r1 <= tdm_band_end)

            in_auto = (auto_count > 0) and (r2 >= auto_band_start and r1 <= auto_band_end)

            
            

            if intersects_cols and (in_tdm or in_auto):
                coord = mcr.coord  # e.g., "A9:L9"
                try:
                    ws_btn.unmerge_cells(coord)
                except Exception:
                    # Already unmerged or invalid range; ignore
                    pass

        # ---------- TDM SECTION (Row 6 onward, contiguous) ----------
        wr = 6
        tdm_sum_k = 0.0
        for r in tdm_rows:
            ws_btn.cell(row=wr, column=1,  value=r.get("Application"))            # A: Application
            ws_btn.cell(row=wr, column=2,  value=r.get("SavingsCategory"))        # B: Savings Category
            ws_btn.cell(row=wr, column=3,  value=float(r.get("N", 0)))            # C: N
            ws_btn.cell(row=wr, column=4,  value=float(r.get("N1", 0)))           # D: N1
            ws_btn.cell(row=wr, column=5,  value=float(r.get("PD1", 0)))          # E: PD1
            ws_btn.cell(row=wr, column=6,  value=float(r.get("C1", 0)))           # F: C1
            ws_btn.cell(row=wr, column=7,  value=float(r.get("N2", 0)))           # G: N2
            ws_btn.cell(row=wr, column=8,  value=float(r.get("PD2", 0)))          # H: PD2
            ws_btn.cell(row=wr, column=9,  value=float(r.get("C2", 0)))           # I: C2
            ws_btn.cell(row=wr, column=10, value=float(r.get("N3", 0)))           # J: N3
            val_k = float(r.get("ROI", 0))                                        # K: $ Savings
            ws_btn.cell(row=wr, column=11, value=val_k)
            ws_btn.cell(row=wr, column=12, value=float(r.get("SavingsPD", 0)))    # L: Savings (PDs)
            tdm_sum_k += val_k
            wr += 1

        # ---------- AUTOMATION SECTION ----------
        # Start AFTER a fixed spacer that exists in the template: 9 rows below row 6,
        # plus however many TDM rows were written. No row insertions.
        y  = 10 + len(tdm_rows)
        wr = y
        auto_sum_k = 0.0
        for r in auto_rows:
            ws_btn.cell(row=wr, column=1,  value=r.get("Application"))            # A: Application
            ws_btn.cell(row=wr, column=2,  value=r.get("SavingsCategory"))        # B: Savings Category
            ws_btn.cell(row=wr, column=3,  value=float(r.get("N", 0)))            # C: N
            ws_btn.cell(row=wr, column=4,  value=float(r.get("N1", 0)))           # D: N1
            ws_btn.cell(row=wr, column=5,  value=float(r.get("PD1", 0)))          # E: PD1
            ws_btn.cell(row=wr, column=6,  value=float(r.get("C1", 0)))           # F: C1
            ws_btn.cell(row=wr, column=7,  value=float(r.get("N2", 0)))           # G: N2
            ws_btn.cell(row=wr, column=8,  value=float(r.get("PD2", 0)))          # H: PD2
            ws_btn.cell(row=wr, column=9,  value=float(r.get("C2", 0)))           # I: C2
            ws_btn.cell(row=wr, column=10, value=float(r.get("N3", 0)))           # J: N3
            val_k = float(r.get("ROI", 0))                                        # K: $ Savings
            ws_btn.cell(row=wr, column=11, value=val_k)
            ws_btn.cell(row=wr, column=12, value=float(r.get("SavingsPD", 0)))    # L: Savings (PDs)
            auto_sum_k += val_k
            wr += 1

       # FIX: Ensure the Row 16 header merges (D–F and G–I) exist (for all *button* sheets),
       # including "Standalone Robot" and "Robot (With Selenium Playwright)".
        
        '''
        try:
           ws_btn.merge_cells(start_row=16, start_column=4, end_row=16, end_column=6)  # D16:F16
           ws_btn.merge_cells(start_row=16, start_column=7, end_row=16, end_column=9)  # G16:I16
        except Exception:
           pass
       # Populate header text if template text was lost/blank

        #if not (ws_btn.cell(row=16, column=4).value or "").strip():
        _v_df = ws_btn.cell(row=16, column=4).value
        if not (isinstance(_v_df, str) and _v_df.strip()):
           ws_btn.cell(row=16, column=4, value="Pre-Capability [eg. Manual]")  # FIX

        #if not (ws_btn.cell(row=16, column=7).value or "").strip():
        _v_gi = ws_btn.cell(row=16, column=7).value
        if not (isinstance(_v_gi, str) and _v_gi.strip()):
            ws_btn.cell(row=16, column=7, value="Post-Capability Automation")   # FIX
        '''
        # ---------- Totals (use explicit sums we tracked; avoid touching template formulas/merges) ----------
        #ws_btn["AE9"].value = round(tdm_sum_k, 2)
        #ws_btn["E9"].value  = round(auto_sum_k, 2)
        #ws_btn["D9"].value  = round(tdm_sum_k + auto_sum_k, 2)
        #ws_btn["E1"].value  = round(tdm_sum_k, 2)

        try:
            ws_btn["E1"].value = round(tdm_sum_k, 2)
        except Exception:
            pass
        
        e6_row = 5 + tdm_count
        ws_btn[f"E{e6_row}"].value = round(auto_sum_k, 2)

       # FIX [2]: Update the *shifted* D12 cell to be the sum of E1 and shifted E6.
       # In the default template it's D12; after insertion it moves to 12 + tdm_count.
        d12_row = 12 + tdm_count + auto_count
       
       # Use the values we just wrote (or 0.0 if missing) to mirror "E1 + E6" explicitly.
        e1_val = ws_btn["E1"].value or 0.0
        e6_val = ws_btn[f"E{e6_row}"].value or 0.0

        ws_btn[f"D{d12_row}"].value = round(float(e1_val) + float(e6_val), 2)

        cum_cell = ws_btn[f"D{d12_row}"]
        cum_cell.font = Font(bold=True, color="FF006100")  # dark green
        cum_cell.number_format = '"$"#,##0.00'

        # FIX (1b): Put 'Cumulative Savings' in the adjacent left cell (bold)
        label_cell = ws_btn[f"C{d12_row}"]
        label_cell.value = "Cumulative Savings"
        label_cell.font = Font(bold=True)
        
        
        targets = {
            "Automation Savings — Breakdown": 3,
            "Pre-Capability [eg. Manual]": 3,
            "Post-Capability Automation": 4,
        }

        # Helper: check if (r,c) is already inside a merged range covering at least 'span' cells on this row.
        def _already_merged(ws, r, c, span):
            for mcr in ws.merged_cells.ranges:
                if mcr.min_row == r and mcr.max_row == r and mcr.min_col <= c <= mcr.max_col:
                    # existing width of merge on this row
                    if (mcr.max_col - mcr.min_col + 1) >= span:
                        return True
            return False

        max_r = ws_btn.max_row
        max_c = min(ws_btn.max_column, 50)  # we only need to scan the left side of the template
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v = ws_btn.cell(row=r, column=c).value
                if not isinstance(v, str):
                    continue
                v_str = v.strip()
                if v_str in targets:
                    span = targets[v_str]
                    if not _already_merged(ws_btn, r, c, span):
                        try:
                            ws_btn.merge_cells(start_row=r, start_column=c,
                                               end_row=r, end_column=c + span - 1)
                        except Exception:
                            # If merge overlaps an existing region, safely ignore
                            pass
        # ------------------------------------------------------------

       # ------------------------------------------------------------

    # >>> FIX: Replace the whole function with this version (surgical change only)
    def _write_soa_sheet(button_name: str, rows: list):
        """
        Clone the SOA template and populate ONLY the rows that belong to the
        'SOA API' button (AutomationFrmk == 'SOA API' as grouped by the caller).
        Writes starting at row 3 with mapping:
        B: Application
        C: SOA Package Status
        E: Total Unique Forms
        F: N (Total Automated Test Case Count)
        G: Forms Complexity
        H: Automated Test Case Execution Count (WITH SOA)
        I: Manual Test Case Execution Count PD (Without SOA)
        J: Total PDs of Automation effort - WITH SOA Package
        K: Total PDs of Manual effort - Without SOA Package
        L: Savings (in PDs)
        M: $ Savings [= (C1 - C2) * N3]
        """
        # Use the SOA template; keep sheet title = button name
        ws_btn = _copy_template(ws_soa_template, button_name or "SOA API")
        if ws_btn is None:
            return

        # >>> FIX: Do NOT hard-filter to specific "automation" category names (that caused empty output).
        # Take all rows under this button and only exclude well-known TDM buckets.
        def _is_tdm(cat: str) -> bool:
            return (cat or "").strip() in tdm_cats  # tdm_cats is defined above

        soa_rows = [r for r in rows if not _is_tdm(r.get("SavingsCategory"))]

        if not soa_rows:
            # Nothing to write for this month/button; leave the cloned sheet as-is.
            return

        start_row = 3
        if len(soa_rows) > 1:
            ws_btn.insert_rows(idx=start_row + 1, amount=len(soa_rows) - 1)

        def _get(r, *keys, default=""):
            for k in keys:
                v = r.get(k)
                if v not in (None, ""):
                    return v
            return default

        def _num(v, default=0.0):
            try:
                return float(v)
            except (TypeError, ValueError):
                return default

        # >>> FIX: Write mapped columns B..M for each row (iterate all UI rows for the selected month)
        for i, r in enumerate(soa_rows):
            rr = start_row + i

            # B: Application
            ws_btn.cell(row=rr, column=2, value=_get(r, "Application"))

            # C: SOA Package Status (prefer explicit; fallback to SavingsCategory)
            ws_btn.cell(row=rr, column=3, value=_get(r, "SOAPackageStatus", "SOA_Package_Status", "SavingsCategory"))

            # E: Total Unique Forms (several pipeline variants)
            ws_btn.cell(row=rr, column=5, value=_get(r, "TotalUniqueForms", "TotalUniqueFormsDesigned", "TotalUniqueFormsExecuted", default=None))

            # F: N (Total Automated Test Case Count)
            ws_btn.cell(row=rr, column=6, value=_num(_get(r, "N", default=0)))

            # G: Forms Complexity
            ws_btn.cell(row=rr, column=7, value=_get(r, "FormsComplexity"))

            # H: Automated Test Case Execution Count (WITH SOA) → N2
            ws_btn.cell(row=rr, column=8, value=_num(_get(r, "N2", default=0)))

            # I: Manual Test Case Execution Count PD (Without SOA) → N1
            ws_btn.cell(row=rr, column=9, value=_num(_get(r, "N1", default=0)))

            # J: Total PDs of Automation effort - WITH SOA Package → PD2
            ws_btn.cell(row=rr, column=10, value=_num(_get(r, "PD2", default=0)))

            # K: Total PDs of Manual effort - Without SOA Package → PD1
            ws_btn.cell(row=rr, column=11, value=_num(_get(r, "PD1", default=0)))

            # L: Savings (in PDs)
            ws_btn.cell(row=rr, column=12, value=_num(_get(r, "SavingsPD", default=0)))

            # M: $ Savings
            ws_btn.cell(row=rr, column=13, value=_num(_get(r, "ROI", default=0)))


    # Create one sheet per "button" (AutomationFrmk); SOA API uses 'SOA' template; others use 'NonSOA'
    for _button, _rows in _by_button.items():
        name_up = (_button or "").strip().upper()
        if name_up in ("SOA", "SOA API", "API", "SOA-API"):
            _write_soa_sheet(_button, _rows)
        else:
            _write_nonsoa_sheet(_button, _rows)

    conn.close()

    # Stream & name
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    short_name = (intake_name[:10] or "Project").strip()
    filename = f"ROI_{intake_number}-{short_name} - {roi_month}_v1.0.xlsx"
    return bio, filename


