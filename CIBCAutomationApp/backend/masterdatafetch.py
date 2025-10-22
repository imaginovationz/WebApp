# backend/masterdatafetch.py

import os
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# You can place Metric.xlsx in backend/ (same folder as app.py)
METRIC_XLSX = os.path.join(BASE_DIR, "Metric.xlsx")

def _safe_cell(v):
    return "" if v is None else str(v)

def _sheet_to_table(ws):
    """
    Convert a worksheet to {headers, rows} by:
      - finding the first non-empty row as headers
      - reading all subsequent rows of the same width
    """
    headers = []
    rows = []
    first_header_row_idx = None

    # Detect first header row
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [(_safe_cell(x)) for x in row]
        has_any = any([x.strip() for x in vals])
        if not has_any:
            continue
        headers = [h.strip() for h in vals]
        first_header_row_idx = i
        break

    if not headers:
        return {"headers": [], "rows": []}

    ncols = len(headers)

    # Read subsequent rows
    for j, row in enumerate(ws.iter_rows(min_row=first_header_row_idx + 1, values_only=True), start=first_header_row_idx + 1):
        vals = [(_safe_cell(x)) for x in row]
        if not any([x.strip() for x in vals]):
            continue  # skip blank rows
        # pad/trim to header width
        if len(vals) < ncols:
            vals = vals + [""] * (ncols - len(vals))
        elif len(vals) > ncols:
            vals = vals[:ncols]
        rows.append(vals)

    return {"headers": headers, "rows": rows}

def _load_sheet(sheet_name):
    wb = load_workbook(METRIC_XLSX, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {"headers": [], "rows": []}
    ws = wb[sheet_name]
    return _sheet_to_table(ws)

def get_sheet_table_consolidated(tab: str):
    """
    tab: 'creation' | 'execution'
    creation  -> sheet 'Creation'
    execution -> sheet 'Execution' (if absent, returns empty table)
    """
    sheet = "Creation" if tab == "creation" else "Execution"
    return _load_sheet(sheet)

def get_sheet_table_monthly(tab: str, fy: str = "", breakdown: str = ""):
    """
    tab: 'creation' | 'execution'
    creation  -> sheet 'MonthlyCreation'
    execution -> sheet 'MonthlyExecution'
    Optional 'fy' filtering:
      - If sheet has a header that looks like 'FY' or 'Year' we filter equals (except 'All Years').
      - If not found, we return entire sheet (frontend shows full table).
    'breakdown' is a no-op placeholder for now; can be used later to pivot/group.
    """
    sheet = "MonthlyCreation" if tab == "creation" else "MonthlyExecution"
    table = _load_sheet(sheet)

    # Try a simple filter if the first column is FY-like
    if fy and fy.lower() != "all years" and table["headers"]:
        headers = table["headers"]
        rows = table["rows"]
        # find a column that looks like FY
        candidate_cols = [i for i, h in enumerate(headers) if h.strip().lower() in ("fy", "year", "fiscal year")]
        if candidate_cols:
            col = candidate_cols[0]
            filtered = [r for r in rows if (r[col] or "").strip().lower() == fy.strip().lower()]
            return {"headers": headers, "rows": filtered}
    return table


