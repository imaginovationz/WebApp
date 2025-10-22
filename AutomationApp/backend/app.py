# backend/app.py
from flask_cors import CORS
import sqlite3
from datetime import datetime
import pbi_dynamic
from flask import Flask, request, jsonify, send_file   # CHANGE
from io import BytesIO                                 # CHANGE
from openpyxl import Workbook                          # CHANGE
from openpyxl.styles import Font, PatternFill          # CHANGE
TEMPLATE_XLSX = "ROI_Portfolio_Project Intake number and name_version.xlsx"
import os  # NEW
from ROIExcelDownload import build_roi_workbook  # NEW
import base64
from datetime import date, datetime
import threading
_db_write_lock = threading.Lock()


from masterdatafetch import (
    get_sheet_table_consolidated,
    get_sheet_table_monthly
)


# [EXCEL-INTEGRATION] --- begin
from flask import send_from_directory, abort
from urllib.parse import unquote
import os
from werkzeug.utils import secure_filename   # if not already imported

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

AUTOMATION_SQL_FILE = os.path.join(BASE_DIR, "AutomationQueries.sql")
ALLOWED_EXTS = {".xlsx", ".xlsm"}

def sanitize_filename_exact(name: str) -> str:
    """
    Keep spaces and hyphens intact; block traversal and disallowed extensions.
    """
    # remove any directory components
    base = os.path.basename(name)
    # forbid slashes/backslashes just in case
    if "/" in base or "\\" in base:
        abort(400, "Invalid filename")
    # enforce extension
    ext = os.path.splitext(base)[1].lower()
    if ext not in ALLOWED_EXTS:
        abort(400, "Only .xlsx allowed")
    # optional: trim control chars
    base = "".join(ch for ch in base if ch >= " " )
    if not base:
        abort(400, "Empty filename")
    return base


def _make_json_safe(obj):
    """Recursively convert bytes/datetime to JSON-serializable values."""
    if isinstance(obj, dict):
        return {k: _make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_make_json_safe(v) for v in obj]
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    if isinstance(obj, bytes):
        # Try to decode text; if not text, return base64 so client can handle it
        try:
            return obj.decode("utf-8")
        except Exception:
            return base64.b64encode(obj).decode("ascii")
    return obj





#for SP list
from HarvestHrSP import fetch_harvested_hours

# --- Optional: if these modules aren't present in your env, the try/except keeps the app running
try:
    from SQL_Initiatives import (
        GET_ALL_OPTIONS, GET_OPTION_BY_ID, UPDATE_OPTION,
        GET_ALL_INITIATIVES, UPDATE_INITIATIVE
    )
except Exception:
    # Fallback SQL (safe defaults if SQL_Initiatives module isn't available)
    GET_ALL_OPTIONS = "SELECT id, options FROM options"
    GET_OPTION_BY_ID = "SELECT id, options FROM options WHERE id = ?"
    UPDATE_OPTION = "UPDATE options SET options = ? WHERE id = ? AND options = ?"
    GET_ALL_INITIATIVES = "SELECT * FROM QEInitiatives_Summary"
    # WARNING: placeholder; adjust to your schema if you rely on this
    UPDATE_INITIATIVE = """
        UPDATE QEInitiatives_Summary
        SET IniName=?, InitiativeDescription=?, InitiativeStatus=?, InitiativeCommentary=?, CumulativeROI=?
        WHERE InitiativeID = ?
    """

# Optional ALM connector: safely stubbed if not present
try:
    from ALMConnector import fetch_tc_count  # noqa
except Exception:
    def fetch_tc_count(query: str):
        return {"count": 0, "query": query}

app = Flask(__name__)
CORS(app)

# ---------------------------------------------------------------------
# DB helper
# ---------------------------------------------------------------------
def get_db_connection():
    #conn = sqlite3.connect('database.db')
    #conn.row_factory = sqlite3.Row
    #return conn

    conn = sqlite3.connect('database.db', timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    # Make SQLite more concurrent-friendly
    conn.execute("PRAGMA journal_mode=WAL;")  #allows reads and writes to happen concurrently.
    conn.execute("PRAGMA busy_timeout=5000;")      # wait up to 5s instead of failing
    conn.execute("PRAGMA synchronous=NORMAL;")     # good default for WAL
    return conn


def _read_sql_section(filepath: str, key: str) -> str:
    """
    Reads a named query from the .sql file.
    Use markers in the file like:
        -- QUERY: refresh_master_inventory
        ... SQL ...
        -- END
    If the marker isn't found and key == 'refresh_master_inventory',
    we fallback to executing the entire file (back-compat with current file).
    """
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except Exception as e:
        raise RuntimeError(f"Failed to read SQL file: {e}")

    start_tag = f"-- QUERY: {key}"
    end_tag = "-- END"

    start_idx, end_idx = None, None
    for i, line in enumerate(lines):
        if line.strip().upper() == start_tag.upper():
            start_idx = i + 1
            break
    if start_idx is not None:
        for j in range(start_idx, len(lines)):
            if lines[j].strip().upper() == end_tag:
                end_idx = j
                break
        section = "".join(lines[start_idx:end_idx or len(lines)])
        if not section.strip():
            raise RuntimeError(f"Query '{key}' has no content")
        return section

    # fallback: entire file for initial single-query file
    if key == "refresh_master_inventory":
        return "".join(lines)

    raise RuntimeError(f"Query '{key}' not found in {os.path.basename(filepath)}")



# [EXCEL-INTEGRATION] --- begin
@app.route("/api/upload_xlsx", methods=["POST"])
def upload_xlsx():
    if "file" not in request.files:
        return "No file", 400
    f = request.files["file"]

    # keep the original client name (with spaces), safely
    original = f.filename or "workbook.xlsx"
    try:
        original = unquote(original)
    except Exception:
        pass
    filename = sanitize_filename_exact(original)

    path = os.path.join(UPLOAD_DIR, filename)
    f.save(path)
    return jsonify({"saved_as": filename})


@app.route("/api/files/<path:filename>", methods=["GET"])
def get_file(filename):
    decoded = unquote(filename)
    path = os.path.join(UPLOAD_DIR, decoded)
    if os.path.isfile(path):
        return send_from_directory(UPLOAD_DIR, decoded)
    abort(404)




# ---------------------------------------------------------------------
# Options (kept from your original app)
# ---------------------------------------------------------------------
@app.route('/api/options', methods=['GET'])
def get_options():
    conn = get_db_connection()
    rows = conn.execute(GET_ALL_OPTIONS).fetchall()
    
    conn.close()
    return jsonify([{'id': row['id'], 'options': row['options']} for row in rows])



@app.route('/api/details/<int:option_id>', methods=['GET'])
def get_option_details(option_id):
    conn = get_db_connection()
    row = conn.execute(GET_OPTION_BY_ID, (option_id,)).fetchone()
    conn.close()
    if row:
        return jsonify({'id': row['id'], 'options': row['options']})
    return jsonify({'error': 'Option not found'}), 404


@app.route('/api/update', methods=['POST'])
def update_option():
    data = request.json
    option_id = data.get('id')
    old_value = data.get('old_value')
    new_value = data.get('new_value')

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(UPDATE_OPTION, (new_value, option_id, old_value))
    conn.commit()
    updated = cursor.rowcount
    cursor.close()
    conn.close()

    if updated:
        return jsonify({'status': 'success', 'message': 'Value updated'})
    return jsonify({'status': 'fail', 'message': 'No matching record found'}), 400


# ---------------------------------------------------------------------
# Initiatives (kept; adjust if you rely on UPDATE_INITIATIVE semantics)
# ---------------------------------------------------------------------
@app.route('/api/initiatives', methods=['GET'])
def getInitiative_options():
    try:
        conn = get_db_connection()
        rows = conn.execute(GET_ALL_INITIATIVES).fetchall()
        conn.close()

        if not rows:
            return jsonify({'initiatives': []}), 200

        available_keys = rows[0].keys()
        initiatives = []
        for row in rows:
            initiatives.append({
                "InitiativeName": row["IniName"] if "IniName" in available_keys else "",
                "InitiativeDescription": row["InitiativeDescription"] if "InitiativeDescription" in available_keys else "",
                "InitiativeStatus": row["InitiativeStatus"] if "InitiativeStatus" in available_keys else "",
                "InitiativeCommentary": row["InitiativeCommentary"] if "InitiativeCommentary" in available_keys else "",
                "InitiativeID": row["InitiativeID"] if "InitiativeID" in available_keys else None,
                "CumulativeROI": row["CumulativeROI"] if "CumulativeROI" in available_keys else ""
            })
        return jsonify({"initiatives": initiatives})
    except Exception as e:
        print(f"Error fetching initiatives: {e}")
        return jsonify({'error': 'Failed to fetch initiatives'}), 500


@app.route('/api/updateInitiatives', methods=['POST'])
def updateInitiative_options():
    data = request.json.get('initiatives', [])
    conn = get_db_connection()
    cursor = conn.cursor()

    for initiative in data:
        cursor.execute(UPDATE_INITIATIVE, (
            initiative.get('IniName'),
            initiative.get('InitiativeDescription'),
            initiative.get('InitiativeStatus'),
            initiative.get('InitiativeCommentary'),
            initiative.get('CumulativeROI'),
            initiative.get('InitiativeID'),
        ))

    conn.commit()
    cursor.close()
    conn.close()
    
    return jsonify({'status': 'success', 'message': 'Initiatives updated successfully'})


# ---------------------------------------------------------------------
# Project ROI
# ---------------------------------------------------------------------
@app.route('/api/projectroi/<int:intake_number>', methods=['GET'])
def getProjectROI(intake_number):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        rows = cursor.execute("SELECT * FROM projectroi WHERE intake_number = ?", (intake_number,)).fetchall()
        cursor.close()
        conn.close()
        if not rows:
            return jsonify({'message': 'No ROI data found for this project'}), 404
        return jsonify([dict(row) for row in rows])
    except Exception as e:
        print(f"Error fetching ROI: {e}")
        return jsonify({'error': 'Failed to fetch ROI data'}), 500


@app.route('/api/projectroiupdate', methods=['POST'])
def updateProjectROI():
    try:
        
        
        payload = request.get_json(force=True) or {}
        # CHANGE: support bulk rows from TDM => payload.entries = [{...}, ...]
        entries = payload.get("entries")
        if not entries:
            # Back-compat: old single-insert path (unchanged)
            data = payload
            intake_number = data.get('intake_number')
            conn = get_db_connection()
            cursor = conn.cursor()
            insert_query = """
               INSERT INTO projectroi (
                    FY, ROIReportingQr, ROIReportingMonth, Portfolio, Application, Amount,
                    AutomationFramework, TransformationInitiative, FP, DealCount,
                    TestCaseDesign, TestUpdated, TestCaseExecuted, ROISheet, Lead, Comment, intake_number
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            cursor.execute(insert_query, (
                data.get('FY'), data.get('ROIReportingQr'), data.get('ROIReportingMonth'),
                data.get('Portfolio'), data.get('Application'), data.get('Amount'),
                data.get('AutomationFramework'), data.get('TransformationInitiative'),
                data.get('FP'), data.get('DealCount'), data.get('TestCaseDesign'),
                data.get('TestUpdated'), data.get('TestCaseExecuted'),
                data.get('ROISheet'), data.get('Lead'), data.get('Comment'),
               intake_number
            ))
            conn.commit()
            ok = cursor.rowcount > 0
            cursor.close()
            conn.close()
            return jsonify({'status': 'success' if ok else 'fail'})

       # ---- New bulk insert path (TDM) ----
        conn = get_db_connection()
        cur = conn.cursor()

        # columns to allow both old/new schemas
        cur.execute("PRAGMA table_info(projectroi)")
        colset = {r[1] for r in cur.fetchall()}

        def pick(mapping, src):
            """Return a dict of {db_col: value} where db_col exists and src has the logical value."""
            out = {}
            for logical_key, db_cols in mapping.items():
                val = src.get(logical_key)
                if val is None:
                   continue
                # db_cols is list of fallbacks; use the first that exists in table
                for c in db_cols:
                    if c in colset:
                        out[c] = val
                        break
            return out

        inserted = 0
        for row in entries:
            # Map logical keys -> possible physical columns (prefer the "new" names first)
            mapping = {
               "ROI_FY":            ["ROI_FY", "FY"],
                "ROI_QR":            ["ROI_QR", "ROIReportingQr"],
                "ROI_month":         ["ROI_month", "ROIReportingMonth"],
                "Domain":            ["Domain", "Portfolio"],
                "Application":       ["Application"],
                "ROI":               ["ROI", "Amount"],
                "AutomationFrmk":    ["AutomationFrmk", "AutomationFramework"],
                "SavingsCategory":   ["SavingsCategory", "TransformationInitiative"],
                "TranInitiative":    ["TranInitiative", "TransformationInitiative"],
               "FPItem":            ["FPItem", "FP"],
                "ROISheetLocation":  ["ROISheetLocation", "ROISheet"],
                "QELead":            ["QELead", "Lead"],
                "Comment":           ["Comment"],
                "intake_name":       ["intake_name", "ProjectName"],
                "intake_number":     ["intake_number"],
                "N1_AvgManualPD":                 ["N1_AvgManualPD"],
                "PD1_TotalManualPDsNByN1":        ["PD1_TotalManualPDsNByN1","PD1"],
                "C1_TotalManualCost_D":           ["C1_TotalManualCost_D"],
                "N2_AvgAutomationPD":             ["N2_AvgAutomationPD"],
                "PD2_TotalAutomatedPDsNByN2":     ["PD2_TotalAutomatedPDsNByN2"],
               "C2_TotalAutomationCost_D":       ["C2_TotalAutomationCost_D"],
                "N3_Numberofcycles":              ["N3_Numberofcycles"],
                "SavingsPD":                      ["SavingsPD"],

            }
            to_insert = pick(mapping, row)
            if not to_insert:
                continue
            cols = list(to_insert.keys())
            vals = [to_insert[c] for c in cols]
            placeholders = ",".join(["?"] * len(cols))
            cur.execute(f"INSERT INTO projectroi ({','.join(cols)}) VALUES ({placeholders})", vals)
            inserted += (1 if cur.rowcount > 0 else 0)

        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'inserted': inserted})
   
   
    
    
    
    
    except Exception as e:
        print(f"Error updating ROI: {e}")
        return jsonify({'error': 'Failed to update ROI data'}), 500



# fetch automated utilities for dropdowns
@app.route('/api/automatedutilities', methods=['GET'])
def fetch_automated_utilities():
    """
   Query params:
      - category=TDM        -> return only TDM utilities
      - category=NON_TDM    -> return utilities where Category is NOT 'TDM'
      - (absent/other)      -> return all
    Response: { "utilities": ["Utility A", "Utility B", ...] }
    """
    try:
        cat = (request.args.get('category') or '').strip().lower()
        conn = get_db_connection()
        cur = conn.cursor()
        if cat == 'tdm':
            cur.execute(
                "SELECT UtilityName FROM automatedutilities WHERE Category = 'TDM' ORDER BY UtilityName COLLATE NOCASE;"
            )
        elif cat in ('non_tdm', 'not_tdm', 'not-tdm'):
            cur.execute(
                "SELECT UtilityName FROM automatedutilities WHERE (Category IS NULL OR Category <> 'TDM') ORDER BY UtilityName COLLATE NOCASE;"
            )
        else:
            cur.execute(
                "SELECT UtilityName FROM automatedutilities ORDER BY UtilityName COLLATE NOCASE;"
           )
        rows = cur.fetchall()
        conn.close()
        utilities = [r['UtilityName'] if isinstance(r, sqlite3.Row) else r[0] for r in rows]
        return jsonify({"utilities": utilities})
    except Exception as e:
        print("automatedutilities error:", e)
        return jsonify({"utilities": []}), 200

#---------------------------------------------------------------------
# Project ROI Entry (save grid rows)
# ---------------------------------------------------------------------
@app.route('/api/projectroientry', methods=['POST'])
def save_project_roi_entry():
    try:
        data = request.json.get('entries', [])
        if not data:
            return jsonify({'status': 'fail', 'message': 'No data provided'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        insert_query = """
            INSERT INTO ProjectROIEntry (
                intake_number, Release, ROIMonth, SavingsCategory, AutomationFmk,
                TotalTCsCount, ManualTCsPD, AutomatedTCCreatedPD, NumberofCycles,
                ProjectName, AutomationLead
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        for entry in data:
            cursor.execute(insert_query, (
                entry.get('intake_number'),
                entry.get('Release'),
                entry.get('ROIMonth'),
                entry.get('SavingsCategory'),
                entry.get('AutomationFmk'),
                entry.get('TotalTCsCount'),
                entry.get('ManualTCsPD'),
                entry.get('AutomatedTCCreatedPD'),
                entry.get('NumberofCycles'),
                entry.get('ProjectName'),
                entry.get('AutomationLead')
            ))

        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Entries saved successfully'}), 201
    except Exception as e:
        print(f"Error saving Project ROI Entry: {e}")
        return jsonify({'error': 'Failed to save entries'}), 500


# ---------------------------------------------------------------------
# ALM connector passthrough (safe stub if ALMConnector missing)
# ---------------------------------------------------------------------
@app.route('/api/fetch_tc_count', methods=['GET'])
def get_tc_count():
    query = request.args.get("query", "")
    result = fetch_tc_count(query)
    return jsonify(result)




@app.route('/api/projecttcountupdate', methods=['POST'])
def projecttcountupdate():

    """
    try:
     payload = request.get_json(force=True) or {}
     entries = payload.get("entries", [])               # TDM (existing)
     design_entries = payload.get("design_entries", []) # CHANGE: new
     execution_entries = payload.get("execution_entries", [])  # CHANGE: new
     if not (isinstance(entries, list) and isinstance(design_entries, list) and isinstance(execution_entries, list)):
         return jsonify({"error": "Invalid payload"}), 400
     if not entries and not design_entries and not execution_entries:
         return jsonify({"error": "No entries to insert"}), 400  

     now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
     conn = get_db_connection()
     cur = conn.cursor()

      # Ensure required columns exist (no-op if already present)
     try:
         cur.execute("PRAGMA table_info(projecttccount_TDM)")
         tdm_cols = {r[1] for r in cur.fetchall()}

         required = {"project","Domain","AutomationFrmk","application","LastUpdate","AutoTDMCountExec","DSRRef"}
         missing = required - tdm_cols
         
          # If columns are missing we don't mutate schema silently; just proceed with intersection
     except Exception:
          cols = set()

     inserted = 0
     inserted_design = 0       
     inserted_execution = 0     

     for row in entries:
        project          = row.get("project","").strip()
        domain           = row.get("Domain","").strip()
        frmk             = row.get("AutomationFrmk","").strip()
        application      = row.get("application","").strip()
        auto_count_exec  = row.get("AutoTDMCountExec", 0)
        dsr_ref          = row.get("DSRRef","").strip()
        comment          = row.get("Comment","").strip() 

        insert_cols = ["project","Domain","AutomationFrmk","application","LastUpdate","AutoTDMCountExec","DSRRef","Comment"]
        insert_vals = [project, domain, frmk, application, now_ts, auto_count_exec, dsr_ref, comment]
        sc = (row.get("SavingsCategory") or "").strip()
        if "SavingsCategory" in tdm_cols:
            insert_cols.append("SavingsCategory")
            insert_vals.append(sc)
        placeholders = ",".join(["?"]*len(insert_cols))
        cur.execute(f"INSERT INTO projecttccount_TDM ({','.join(insert_cols)}) VALUES ({placeholders})", insert_vals)


        if cur.rowcount > 0:
            inserted += 1
              
              
         # ======  insert into projecttccount_Design ======
     if design_entries:
         try:
             cur.execute("PRAGMA table_info(projecttccount_Design)")
             design_cols = {r[1] for r in cur.fetchall()}
         except Exception:
             design_cols = set()
         for row in design_entries:
             project     = (row.get("project") or "").strip()
             domain      = (row.get("Domain") or "").strip()
             frmk        = (row.get("AutomationFrmk") or "").strip()
             application = (row.get("application") or "").strip()
             dsr_ref     = (row.get("DSRRef") or "").strip()
             new_cnt     = int(row.get("TotalNewAutoTCCreated") or 0)
             upd_cnt     = int(row.get("TotalExistingAutoTCUpdated") or 0)
             uniq_forms  = row.get("TotalUniqueFormsDesigned")
             man_created = row.get("TotalManualTCCreated")
             sc          = (row.get("SavingsCategory") or "").strip()
             
             # Prefer dynamic insert using available columns; LastUpdate set to now_ts
             cols = ["project","Domain","AutomationFrmk","application","LastUpdate","DSRRef"]
             vals = [project, domain, frmk, application, now_ts, dsr_ref]
             if "TotalNewAutoTCCreated" in design_cols:
                cols.append("TotalNewAutoTCCreated"); vals.append(new_cnt)
             if "TotalExistingAutoTCUpdated" in design_cols:
                cols.append("TotalExistingAutoTCUpdated"); vals.append(upd_cnt)

             #  include new SOA-specific columns when table supports them
             if "TotalUniqueFormsDesigned" in design_cols and uniq_forms is not None:
                cols.append("TotalUniqueFormsDesigned"); vals.append(int(uniq_forms or 0))
             if "TotalManualTCCreated" in design_cols and man_created is not None:
                cols.append("TotalManualTCCreated"); vals.append(int(man_created or 0))
             if "SavingsCategory" in design_cols:                        
                cols.append("SavingsCategory"); vals.append(sc)

             placeholders = ",".join(["?"]*len(cols))
             cur.execute(f"INSERT INTO projecttccount_Design ({','.join(cols)}) VALUES ({placeholders})", vals)
             if cur.rowcount > 0:
                inserted_design += 1

     # ======  insert into projecttccount_Execution ======
     if execution_entries:
         try:
             cur.execute("PRAGMA table_info(projecttccount_Execution)")
             exec_cols = {r[1] for r in cur.fetchall()}
         except Exception:
            exec_cols = set()
         for row in execution_entries:
             project     = (row.get("project") or "").strip()
             domain      = (row.get("Domain") or "").strip()
             frmk        = (row.get("AutomationFrmk") or "").strip()
             application = (row.get("application") or "").strip()
             dsr_ref     = (row.get("DSRRef") or "").strip()
             exec_cnt    = int(row.get("TotalNewAutoTCExecuted") or 0)
             execycle    = (row.get("Execycle") or "").strip()

             uniq_forms  = row.get("TotalUniqueFormsExecuted")
             man_exec    = row.get("TotalManualTCExecuted")
             sc          = (row.get("SavingsCategory") or "").strip()

             cols = ["project","Domain","AutomationFrmk","application","LastUpdate","DSRRef"]
             vals = [project, domain, frmk, application, now_ts, dsr_ref]
             if "TotalNewAutoTCExecuted" in exec_cols:
                 cols.append("TotalNewAutoTCExecuted"); vals.append(exec_cnt)
             if "Execycle" in exec_cols:
                 cols.append("Execycle"); vals.append(execycle)

        #  include new SOA-specific columns when table supports them
             if "TotalUniqueFormsExecuted" in exec_cols and uniq_forms is not None:
                 cols.append("TotalUniqueFormsExecuted"); vals.append(int(uniq_forms or 0))
             if "TotalManualTCExecuted" in exec_cols and man_exec is not None:
                 cols.append("TotalManualTCExecuted"); vals.append(int(man_exec or 0))

             if "SavingsCategory" in exec_cols:                          
                 cols.append("SavingsCategory"); vals.append(sc)

             placeholders = ",".join(["?"]*len(cols))
             cur.execute(f"INSERT INTO projecttccount_Execution ({','.join(cols)}) VALUES ({placeholders})", vals)
             if cur.rowcount > 0:
                 inserted_execution += 1
         

     conn.commit()
     conn.close()
     #return jsonify({"status": "success", "inserted": inserted}), 201
     return jsonify({
         "status": "success",
         "inserted": inserted,                 # TDM rows
         "inserted_design": inserted_design,   # CHANGE
         "inserted_execution": inserted_execution  # CHANGE
     }), 201
    
    
    except Exception as e:
     try:
          conn.rollback()
     except Exception:
          pass
     print("projecttcountupdate error:", e)
     return jsonify({"error": "Failed to insert TDM test-count rows"}), 500

    """

    
    conn = None
    try:
        payload = request.get_json(force=True) or {}
        entries = payload.get("entries", [])                     # TDM section rows
        design_entries = payload.get("design_entries", [])       # Automation (Design categories)
        execution_entries = payload.get("execution_entries", []) # Automation (Execution categories)

        if not (isinstance(entries, list) and isinstance(design_entries, list) and isinstance(execution_entries, list)):
            return jsonify({"error": "Invalid payload"}), 400
        if not entries and not design_entries and not execution_entries:
            return jsonify({"error": "No entries to insert"}), 400

        now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        
        with _db_write_lock:
            conn = get_db_connection()
            cur = conn.cursor()

            inserted = 0
            inserted_design = 0
            inserted_execution = 0

            # ---------- TDM rows -> projecttccount_TDM ----------
            for row in entries:
                project          = (row.get("project") or "").strip()
                domain           = (row.get("Domain") or "").strip()
                frmk             = (row.get("AutomationFrmk") or "").strip()
                application      = (row.get("application") or "").strip()
                auto_count_exec  = int(row.get("AutoTDMCountExec") or 0)
                dsr_ref          = (row.get("DSRRef") or "").strip()
                comment          = (row.get("Comment") or "").strip()
                sc               = (row.get("SavingsCategory") or "").strip()

                cur.execute(
                    """
                    INSERT INTO projecttccount_TDM
                        (project, Domain, AutomationFrmk, application,
                        LastUpdate, AutoTDMCountExec, DSRRef, Comment, SavingsCategory)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (project, domain, frmk, application, now_ts,
                    auto_count_exec, dsr_ref, comment, sc)
                )
                if cur.rowcount > 0:
                    inserted += 1

            # ---------- Automation (Design categories) -> projecttccount_Design ----------
            if design_entries:
                # Keep optional columns dynamic to preserve existing behavior for SOA fields, etc.
                try:
                    cur.execute("PRAGMA table_info(projecttccount_Design)")
                    design_cols = {r[1] for r in cur.fetchall()}
                except Exception:
                    design_cols = set()

                for row in design_entries:
                    project     = (row.get("project") or "").strip()
                    domain      = (row.get("Domain") or "").strip()
                    frmk        = (row.get("AutomationFrmk") or "").strip()
                    application = (row.get("application") or "").strip()
                    dsr_ref     = (row.get("DSRRef") or "").strip()
                    new_cnt     = row.get("TotalNewAutoTCCreated")
                    upd_cnt     = row.get("TotalExistingAutoTCUpdated")
                    uniq_forms  = row.get("TotalUniqueFormsDesigned")
                    man_created = row.get("TotalManualTCCreated")
                    sc          = (row.get("SavingsCategory") or "").strip()

                    cols = ["project","Domain","AutomationFrmk","application","LastUpdate","DSRRef","SavingsCategory"]
                    vals = [project, domain, frmk, application, now_ts, dsr_ref, sc]
                    if "FormsComplexity" in design_cols and row.get("FormsComplexity") is not None:
                        cols.append("FormsComplexity"); vals.append(row.get("FormsComplexity"))

                    if "TotalNewAutoTCCreated" in design_cols and new_cnt is not None:
                        cols.append("TotalNewAutoTCCreated"); vals.append(int(new_cnt or 0))
                    if "TotalExistingAutoTCUpdated" in design_cols and upd_cnt is not None:
                        cols.append("TotalExistingAutoTCUpdated"); vals.append(int(upd_cnt or 0))

                    # SOA-specific optional fields (only if present in schema)
                    if "TotalUniqueFormsDesigned" in design_cols and uniq_forms is not None:
                        cols.append("TotalUniqueFormsDesigned"); vals.append(int(uniq_forms or 0))
                    if "TotalManualTCCreated" in design_cols and man_created is not None:
                        cols.append("TotalManualTCCreated"); vals.append(int(man_created or 0))

                    placeholders = ",".join(["?"] * len(cols))
                    cur.execute(
                        f"INSERT INTO projecttccount_Design ({','.join(cols)}) VALUES ({placeholders})",
                        vals
                    )
                    if cur.rowcount > 0:
                        inserted_design += 1

            # ---------- Automation (Execution categories) -> projecttccount_Execution ----------
            if execution_entries:
                # Keep optional columns dynamic to preserve existing behavior for SOA fields, etc.
                try:
                    cur.execute("PRAGMA table_info(projecttccount_Execution)")
                    exec_cols = {r[1] for r in cur.fetchall()}
                except Exception:
                    exec_cols = set()

                for row in execution_entries:
                    project     = (row.get("project") or "").strip()
                    domain      = (row.get("Domain") or "").strip()
                    frmk        = (row.get("AutomationFrmk") or "").strip()
                    application = (row.get("application") or "").strip()
                    dsr_ref     = (row.get("DSRRef") or "").strip()
                    exec_cnt    = row.get("TotalNewAutoTCExecuted")
                    execycle    = (row.get("Execycle") or "").strip()
                    uniq_forms  = row.get("TotalUniqueFormsExecuted")
                    man_exec    = row.get("TotalManualTCExecuted")
                    sc          = (row.get("SavingsCategory") or "").strip()

                    cols = ["project","Domain","AutomationFrmk","application","LastUpdate","DSRRef","SavingsCategory"]
                    vals = [project, domain, frmk, application, now_ts, dsr_ref, sc]
                    if "FormsComplexity" in exec_cols and row.get("FormsComplexity") is not None:
                        cols.append("FormsComplexity"); vals.append(row.get("FormsComplexity"))
            
                    if "TotalNewAutoTCExecuted" in exec_cols and exec_cnt is not None:
                        cols.append("TotalNewAutoTCExecuted"); vals.append(int(exec_cnt or 0))
                    if "Execycle" in exec_cols and execycle:
                        cols.append("Execycle"); vals.append(execycle)

                    # SOA-specific optional fields (only if present in schema)
                    if "TotalUniqueFormsExecuted" in exec_cols and uniq_forms is not None:
                        cols.append("TotalUniqueFormsExecuted"); vals.append(int(uniq_forms or 0))
                    if "TotalManualTCExecuted" in exec_cols and man_exec is not None:
                        cols.append("TotalManualTCExecuted"); vals.append(int(man_exec or 0))

                    placeholders = ",".join(["?"] * len(cols))
                    cur.execute(
                        f"INSERT INTO projecttccount_Execution ({','.join(cols)}) VALUES ({placeholders})",
                        vals
                    )
                    if cur.rowcount > 0:
                        inserted_execution += 1

            conn.commit()
            conn.close()
            return jsonify({
                "status": "success",
                "inserted": inserted,                     # TDM rows
                "inserted_design": inserted_design,       # Design rows
                "inserted_execution": inserted_execution  # Execution rows
            }), 201

    except Exception as e:
        try:
            if conn:
                conn.rollback()
                conn.close()
        except Exception:
            pass
        print("projecttcountupdate error:", e)
        return jsonify({"error": "Failed to insert TDM test-count rows"}), 500



@app.route("/api/powerbi/filters", methods=["GET"])
def powerbi_filters():
    try:
        m = float(request.args.get("manual_cost"))
        a = float(request.args.get("automation_cost"))
    except Exception:
        return jsonify({"error": "manual_cost and automation_cost are required numeric query params"}), 400
    filters = pbi_dynamic.make_filters(m, a)
    return jsonify({"filters": filters})



# ---------------------------------------------------------------------
# Projects: search + details (USED BY UI)
# ---------------------------------------------------------------------
@app.route('/api/projects/search', methods=['GET'])
def search_projects():
    try:
        q = request.args.get('query', '')
        release = (request.args.get('release') or '').strip()   # CHANGE: optional filter
        limit = int(request.args.get('limit', 20))              # CHANGE: honor limit
        conn = get_db_connection()
        cursor = conn.cursor()

        # CHANGE: if release filter provided, fetch projects for that release (no 3-char requirement)
        if release:
            sql = """
                SELECT intake_number, intake_name
                  FROM projects
                 WHERE release = ?
              GROUP BY intake_number, intake_name
              ORDER BY intake_name COLLATE NOCASE
                 LIMIT ?
            """
            rows = cursor.execute(sql, (release, limit)).fetchall()
            conn.close()
            return jsonify([dict(row) for row in rows])

        # existing behavior for text search:
        if not q or len(q) < 3:
            conn.close()
            return jsonify([])

        #conn = get_db_connection()
        #cursor = conn.cursor()
        sql = """
            SELECT intake_number, intake_name
            FROM projects
            WHERE intake_number LIKE ? OR intake_name LIKE ?
            LIMIT ?
        """
        like_q = f"%{q}%"
        #rows = cursor.execute(sql, (like_q, like_q)).fetchall()
        rows = cursor.execute(sql, (like_q, like_q, limit)).fetchall()
        conn.close()
        return jsonify([dict(row) for row in rows])
    except Exception as e:
        print(f"Error fetching projects: {e}")
        return jsonify({'error': 'Failed to fetch projects'}), 500


@app.route('/api/projects/<string:intake_number>', methods=['GET'])
def get_project_details(intake_number):
    """
    Returns the latest row for this intake_number from `projects`,
    aliases EstiamteAssignedTo -> EstimateAssignedTo,
    and augments ImplementationDate from projects_timeline if available.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        row = cur.execute(
            "SELECT * FROM projects WHERE intake_number = ? ORDER BY timestamp DESC LIMIT 1",
            (intake_number,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'message': 'Project not found'}), 404

        data = dict(row)

        # Alias schema typo: EstiamteAssignedTo -> EstimateAssignedTo
        if 'EstimateAssignedTo' not in data and 'EstiamteAssignedTo' in data:
            data['EstimateAssignedTo'] = data['EstiamteAssignedTo']

        # Optionally include ImplementationDate
        try:
            cur.execute("PRAGMA table_info(projects_timeline)")
            if cur.fetchall():
                r2 = cur.execute(
                    "SELECT ImplementationDate FROM projects_timeline WHERE intake_number = ? ORDER BY ROWID DESC LIMIT 1",
                    (intake_number,)
                ).fetchone()
                if r2:
                    data['ImplementationDate'] = r2['ImplementationDate'] if isinstance(r2, sqlite3.Row) else r2[0]
        except Exception:
            pass

        conn.close()
        return jsonify(data)
    except Exception as e:
        print(f"Error fetching project details: {e}")
        return jsonify({'error': 'Failed to fetch project details'}), 500


# ---------------------------------------------------------------------
# NEW: Automation Deliverables fetch for a project (USED BY ProjectStatusTab)
# ---------------------------------------------------------------------
@app.route('/api/automationdeliverables/<string:intake_number>', methods=['GET'])
def get_automation_deliverables(intake_number):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        row = cur.execute(
            "SELECT * FROM automationdeliverables WHERE intake_number = ? ORDER BY ROWID DESC LIMIT 1",
            (intake_number,)
        ).fetchone()
        conn.close()
        return jsonify(dict(row) if row else {})
    except Exception as e:
        print(f"Error fetching automationdeliverables: {e}")
        return jsonify({'error': 'Failed to fetch automation deliverables'}), 500


# ---------------------------------------------------------------------
# Resource list (dropdowns)
# ---------------------------------------------------------------------
@app.route('/api/resourcefetch', methods=['GET'])
def resourcefetch():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        #cur.execute("SELECT name FROM resources ORDER BY name;")
        #rows = cur.fetchall()
        #conn.close()
        #names = [r['name'] if isinstance(r, sqlite3.Row) else r[0] for r in rows]
        #return jsonify({"resources": names})
        # CHANGE: support lookup by name to fetch email for notifications
        name = request.args.get("name", "").strip()
        if name:
            row = cur.execute("SELECT email FROM resources WHERE name = ? LIMIT 1;", (name,)).fetchone()
            conn.close()
            if not row:
                return jsonify({"error": "Lead not found"}), 404
            email = row["email"] if isinstance(row, sqlite3.Row) else row[0]
            return jsonify({"name": name, "email": email})
        else:
            cur.execute("SELECT name FROM resources ORDER BY name;")
            rows = cur.fetchall()
            conn.close()
            names = [r['name'] if isinstance(r, sqlite3.Row) else r[0] for r in rows]
            return jsonify({"resources": names})
   
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/fetchapplication', methods=['GET'])
def fetch_application():
    """CHANGE: applications list for a given domain (Portfolio)."""
    try:
        domain = (request.args.get("domain") or "").strip()
        if not domain:
           return jsonify({"applications": []})
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
           "SELECT application FROM application WHERE domain = ? ORDER BY application COLLATE NOCASE;",
            (domain,)
        )
        rows = cur.fetchall()
        conn.close()
        apps = [r["application"] if isinstance(r, sqlite3.Row) else r[0] for r in rows]
        return jsonify({"applications": apps})
    except Exception as e:
        print("fetch_application error:", e)
        return jsonify({"applications": []}), 200




@app.route('/api/send_intake_email', methods=['POST'])
def send_intake_email():
    try:
        data = request.get_json(force=True) or {}
        to = data.get("to")
        subject = data.get("subject", "")
        html = data.get("html", "")
        if not to:
            return jsonify({"error": "Missing 'to'"}), 400

        # Import email helper from components/src/Utils
        import os, sys
        base_dir = os.path.dirname(os.path.abspath(__file__))      # .../backend
        repo_root = os.path.dirname(base_dir)                      # project root
        utils_path = os.path.join(repo_root, "frontend", "src", "Utils")
        if utils_path not in sys.path:
           sys.path.append(utils_path)
        from IntakeEmail import send_intake_email as _send_email


        ok, info = _send_email(to, subject, html)
        if ok:
            return jsonify({"ok": True, "info": info})
        return jsonify({"ok": False, "error": info}), 500
    except Exception as e:
        print("send_intake_email error:", e)
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------
# Project Master Update: robust + alias mapping
# ---------------------------------------------------------------------
@app.route('/api/projectmasterupdate', methods=['POST'])
def projectmasterupdate():
    """
    Payload example:
    {
      "intake_number": 12345,
      "updates": [
        {"table":"projects","column":"DeliveryModel","value":"FP"},
        {"table":"projects","column":"manual_cost","value": 100},
        {"table":"projects","column":"automation_cost","value": 40},
        {"table":"projects","column":"EstimateAssignedTo","value":"Alice"},
        {"table":"projects_timeline","column":"ImplementationDate","value":"2025-10-01"},
        {"table":"automationdeliverables","column":"ROIMonth","value":"2025-11"},
        {"table":"InitiativesPlannedUsage","column":"GenAI_PS","value":"YES"}
      ]
    }
    """
    data = request.get_json(force=True)
    intake_number = data.get("intake_number")
    updates = data.get("updates", [])
    if not intake_number or not isinstance(updates, list):
        return jsonify({"error": "Invalid payload"}), 400

    # helper for InitiativesPlannedUsage one-col upserts
    def _upsert_initiatives(conn, intake_number, col, val, intake_name=None):
        cur = conn.cursor()
        # ensure a row exists
        r = cur.execute(
            "SELECT 1 FROM InitiativesPlannedUsage WHERE intake_number = ? LIMIT 1",
            (intake_number,)
        ).fetchone()
        if not r:
            cur.execute(
                "INSERT INTO InitiativesPlannedUsage (intake_number, intake_name) VALUES (?, ?)",
                (intake_number, intake_name or "")
            )
        # update the single column
        cur.execute(f'UPDATE InitiativesPlannedUsage SET "{col}" = ? WHERE intake_number = ?',
                    (val, intake_number))
        conn.commit()

    # Whitelisted columns (extended surgically)
    allowed = {
        "projects": {
            "DeliveryModel","TimeLineAvailable","project_status","QAManager",
            "functional_qe_lead","FuncSME","automation_qe_lead","OffshoreLead",
            "FTE_Architect","ITPM","AutoWorkStartDate","AutoWorkEndDate",
            "AutoEstimateDueDate","AutoEstimateSubmittedOn","AutoEstimateStatus",
            "EstimateAssignedTo","EstiamteAssignedTo",
            "EstimateReviewedBy","FuncAutomationScope","AutoOOSReason","AutomationOOSReason",
            "manual_cost","automation_cost","qe_cost","PSEMinSavings","OOSRef",
            "AutomationStatus","Comments","timestamp",
            "intake_name","PATAssignee","intake_entry_date",
            "Program","release","application","domain",
            "EstimateApprovedBy",          # keep: approved by
            "change_type"                  # keep: Change Type
        },
        "projects_timeline": {"ImplementationDate"},
        "automationdeliverables": {
            # keep existing checkboxes + fields
            "AutomationTestPlan","Checklist","ALMupdate","GITUpdate","DSR",
            "LESubmitted","RegressionTestLab","DashboardMetric","ReleaseHandover",
            "ProjectClosure","ROIMonth","QETransformed","ShiftLeft",
            "intake_name",
            # === NEW (surgical): mirror-of-projects fields ===
            "Domain",       # should receive Portfolio (projects.domain)
            "ChangeType"    # should receive Change Type (projects.change_type)
        },
        "InitiativesPlannedUsage": {
            "intake_number","intake_name","GenAI_PS","DDGS_SOA3","E2E","CGI EnvoyAPI",
            "QEDevops","ExcaliburLetter","Fireflink","mmtgRegressionPhase2",
            "mmtgTDM","CLASSTDM","Conformiq"
        },
    }

    # UI -> physical column aliases
    alias_to_physical = {
        ("projects", "EstimateAssignedTo"): "EstimateAssignedTo",  # schema typo in DB
    }

    def table_columns(cur, table):
        cur.execute(f"PRAGMA table_info({table})")
        return {row[1] for row in cur.fetchall()}

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Identify a target projects row (latest by timestamp) — existing behavior
        cur.execute(
            "SELECT domain, application FROM projects WHERE intake_number = ? ORDER BY timestamp DESC LIMIT 1",
            (intake_number,)
        )
        proj_key = cur.fetchone()  # may be None

        # Pre-fetch existing columns for each table
        proj_cols = table_columns(cur, "projects")
        ad_cols   = table_columns(cur, "automationdeliverables")
        pt_cols   = table_columns(cur, "projects_timeline")

        # Group normalized updates (robust to extra tables)
        grouped = {"projects": {}, "automationdeliverables": {}, "projects_timeline": {}}
        for u in updates:
            t = u.get("table")
            c = u.get("column")
            v = u.get("value")
            if not t or not c or t not in allowed:
                continue
            phys_col = alias_to_physical.get((t, c), c)
            if c not in allowed[t] and phys_col not in allowed[t]:
                continue
            if t not in grouped:
                grouped[t] = {}
            grouped[t][phys_col] = v

        payload_intake_name = (
            (grouped.get("projects", {}) or {}).get("intake_name")
            or (grouped.get("automationdeliverables", {}) or {}).get("intake_name")
            or (grouped.get("InitiativesPlannedUsage", {}) or {}).get("intake_name")
            or ""
        )
        if not payload_intake_name:
            row = cur.execute(
                "SELECT intake_name FROM projects WHERE intake_number = ? ORDER BY timestamp DESC LIMIT 1",
                (intake_number,)
            ).fetchone()
            payload_intake_name = (row["intake_name"].strip() if row and row["intake_name"] else "")

        composite_timeline_key = f"{intake_number}-{payload_intake_name}" if payload_intake_name else str(intake_number)
        

        results = []

        # ---------- PROJECTS: update/insert (existing composite-key behavior)
        proj_update = {c: v for c, v in grouped.get("projects", {}).items() if c in proj_cols}
        if proj_update:
            supplied_name = grouped["projects"].get("intake_name")
            keyrow = None
            if supplied_name:
                keyrow = cur.execute(
                    "SELECT domain, application FROM projects WHERE intake_number = ? AND intake_name = ? ORDER BY timestamp DESC LIMIT 1",
                    (intake_number, supplied_name)
                ).fetchone()

            if keyrow is None and not grouped.get("projects"):
                ts = (grouped.get("projects") or {}).get("timestamp")
                cur.execute(
                    "INSERT INTO projects (intake_number, timestamp) VALUES (?, COALESCE(?, CURRENT_TIMESTAMP))",
                    (intake_number, ts),
                )
                keyrow = (None, None)

            if not keyrow:
                keyrow = proj_key  # may be None
            if keyrow:
                domain = keyrow["domain"] if isinstance(keyrow, sqlite3.Row) else (keyrow[0] if keyrow else None)
                application = keyrow["application"] if isinstance(keyrow, sqlite3.Row) else (keyrow[1] if keyrow else None)

                set_clause = ", ".join([f"{c}=?" for c in proj_update.keys()])
                vals_only = list(proj_update.values())

                cur.execute(
                    f"UPDATE projects SET {set_clause} WHERE intake_number=? AND domain=? AND application=?",
                    vals_only + [intake_number, domain, application]
                )
                if cur.rowcount > 0:
                    results.append({"table": "projects", "status": "updated", "columns": list(proj_update.keys())})
                else:
                    # fallback: (domain, application) empty
                    cur.execute(
                        f"""UPDATE projects
                               SET {set_clause}
                             WHERE intake_number = ?
                               AND (domain IS NULL OR TRIM(domain) = '')
                               AND (application IS NULL OR TRIM(application) = '')""",
                        vals_only + [intake_number]
                    )
                    if cur.rowcount > 0:
                        results.append({"table": "projects", "status": "updated_fallback", "columns": list(proj_update.keys())})
                    else:
                        results.append({"table": "projects", "status": "skipped", "reason": "target row not found for composite or fallback key"})
            else:
                # first time row
                insert_cols = ["intake_number"] + list(proj_update.keys())
                placeholders = ",".join(["?"] * len(insert_cols))
                insert_vals = [intake_number] + list(proj_update.values())
                cur.execute(
                    f"INSERT INTO projects ({','.join(insert_cols)}) VALUES ({placeholders})",
                    insert_vals
                )
                results.append({"table": "projects", "status": "inserted", "columns": list(proj_update.keys())})

        # ---------- mirror Portfolio/Change Type to automationdeliverables
        # Only if the automationdeliverables table actually has these columns.
        if "Domain" in ad_cols and "domain" in grouped.get("projects", {}):
            grouped.setdefault("automationdeliverables", {})
            grouped["automationdeliverables"]["Domain"] = grouped["projects"]["domain"] 
        if "ChangeType" in ad_cols and "change_type" in grouped.get("projects", {}):
            grouped.setdefault("automationdeliverables", {})
            grouped["automationdeliverables"].setdefault("ChangeType", grouped["projects"]["change_type"])
        # Preserve intake_name on automationdeliverables if projects.intake_name provided
        if "intake_name" in ad_cols and "intake_name" in grouped.get("projects", {}):
            grouped.setdefault("automationdeliverables", {})
            grouped["automationdeliverables"].setdefault("intake_name", grouped["projects"]["intake_name"])

        # ---------- AUTOMATIONDELIVERABLES: upsert by intake_number
        ad_update = {c: v for c, v in grouped.get("automationdeliverables", {}).items() if c in ad_cols}
        if ad_update:
            set_clause = ", ".join([f"{c}=?" for c in ad_update.keys()])
            vals = list(ad_update.values()) + [intake_number]
            key_col = "intake_number" if "intake_number" in ad_cols else None
            if key_col:
                cur.execute(f"UPDATE automationdeliverables SET {set_clause} WHERE {key_col}=?", vals)
                if cur.rowcount == 0:
                    cols_with_key = [key_col] + list(ad_update.keys())
                    placeholders = ",".join(["?"] * len(cols_with_key))
                    vals_ins = [intake_number] + list(ad_update.values())
                    cur.execute(
                        f"INSERT INTO automationdeliverables ({','.join(cols_with_key)}) VALUES ({placeholders})",
                        vals_ins
                    )
                    results.append({"table": "automationdeliverables", "status": "inserted", "columns": list(ad_update.keys())})
                else:
                    results.append({"table": "automationdeliverables", "status": "updated", "columns": list(ad_update.keys())})
            else:
                results.append({"table": "automationdeliverables", "status": "skipped", "reason": "no key column"})

        # ---------- PROJECTS_TIMELINE: upsert by intake_number
        pt_update = {c: v for c, v in grouped.get("projects_timeline", {}).items() if c in pt_cols}
        if pt_cols and pt_update:
            key_col = "intake_number" if "intake_number" in pt_cols else None
            set_clause = ", ".join([f"{c}=?" for c in pt_update.keys()])
            if key_col:
                # >>> CHANGED: store "<intake_number>-<intake_name>" in intake_number; migrate existing row if needed
                set_clause_ext = f"{set_clause}, {key_col}=?"
                # Try to UPDATE any row with either the new composite key (idempotent) OR the old numeric key (migration)
                cur.execute(
                    f"UPDATE projects_timeline SET {set_clause_ext} WHERE {key_col} IN (?, ?)",
                    list(pt_update.values()) + [composite_timeline_key, composite_timeline_key, str(intake_number)]
                )
                if cur.rowcount == 0:
                    # If no row existed, INSERT with the composite key
                    cols_with_key = [key_col] + list(pt_update.keys())
                    placeholders = ",".join(["?"] * len(cols_with_key))
                    cur.execute(
                        f"INSERT INTO projects_timeline ({','.join(cols_with_key)}) VALUES ({placeholders})",
                        [composite_timeline_key] + list(pt_update.values())
                    )
                    results.append({"table": "projects_timeline", "status": "inserted", "columns": list(pt_update.keys())})
                else:
                    results.append({"table": "projects_timeline", "status": "updated", "columns": list(pt_update.keys())})
                
            else:
                results.append({"table": "projects_timeline", "status": "skipped", "reason": "no key column"})




        # ---------- INITIATIVES PLANNED USAGE: per-column upsert (existing behavior)
        if "InitiativesPlannedUsage" in grouped and grouped["InitiativesPlannedUsage"]:
            init_payload = grouped["InitiativesPlannedUsage"]
            init_intake_name = init_payload.get("intake_name") or grouped.get("projects", {}).get("intake_name")
            for col, val in init_payload.items():
                _upsert_initiatives(conn, intake_number, col, val, intake_name=init_intake_name)
            results.append({"table": "InitiativesPlannedUsage", "status": "upserted", "columns": list(init_payload.keys())})

        conn.commit()
        conn.close()
        return jsonify({"ok": True, "results": results})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print("projectmasterupdate error:", e)
        return jsonify({"error": str(e)}), 500




@app.route('/api/initiativeplannedusage', methods=['GET'])
def get_initiatives_planned_usage():
    try:
        intake_number = (request.args.get("intake_number") or "").strip()
        if not intake_number:
            return jsonify({}), 200
        conn = get_db_connection()
        cur = conn.cursor()
        row = cur.execute(
            "SELECT * FROM InitiativesPlannedUsage WHERE intake_number = ? ORDER BY ROWID DESC LIMIT 1",
            (intake_number,)
        ).fetchone()
        conn.close()
        return jsonify(dict(row) if row else {})
    except Exception as e:
        print("initiativeplannedusage error:", e)
        return jsonify({}), 200
    




@app.route('/api/fetchreleases', methods=['GET'])
def fetch_releases():
    """CHANGE: provide dropdown values for Release Name."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT release_name FROM releases ORDER BY release_name COLLATE NOCASE;")
        rows = cur.fetchall()
        conn.close()
        releases = [r["release_name"] if isinstance(r, sqlite3.Row) else r[0] for r in rows]
        return jsonify({"releases": releases})
    except Exception as e:
        print("fetch_releases error:", e)
        return jsonify({"releases": []}), 200




# OOS reference upload (saves to projects.OOSRef for latest row)
@app.route('/api/oosref_upload', methods=['POST'])
def oosref_upload():
    try:
        intake_number = request.form.get("intake_number")
        f = request.files.get("file")

        # NEW: read reason + status from the form (names kept flexible)
        # 'reason' maps to "Reason if OOS / No new Func. Automation"
        reason = (
            #request.form.get("reason")
            request.form.get("AutoOOSReason")
            #or request.form.get("AutomationOOSReason")
            #or request.form.get("reason_if_oos")
            #or request.form.get("reasonIfOOS")
        )
        estimate_status = (
            #request.form.get("estimate_status")
            request.form.get("AutoEstimateStatus")
            #or request.form.get("estimateStatusOrOOS")
        )

        if not intake_number or not f:
            return jsonify({"error": "intake_number and file are required"}), 400

        content = f.read()
        conn = get_db_connection()
        cur = conn.cursor()

        # Ensure OOSRef and AutomationOOSReason columns exist
        try:
            cur.execute("PRAGMA table_info(projects)")
            cols = {r[1] for r in cur.fetchall()}
            if "OOSRef" not in cols:
                cur.execute("ALTER TABLE projects ADD COLUMN OOSRef BLOB")
            if "AutomationOOSReason" not in cols:
                cur.execute("ALTER TABLE projects ADD COLUMN AutomationOOSReason TEXT")
        except Exception:
            pass

        # Update latest row for this intake_number (by timestamp)
        if estimate_status == "Automation-OOS" and reason:
            cur.execute("""
                UPDATE projects
                   SET OOSRef = ?, AutomationOOSReason = ?
                 WHERE ROWID = (
                   SELECT ROWID FROM projects
                    WHERE intake_number = ?
                 ORDER BY timestamp DESC LIMIT 1)
            """, (sqlite3.Binary(content), reason, intake_number))
        else:
            cur.execute("""
                UPDATE projects
                   SET OOSRef = ?
                 WHERE ROWID = (
                   SELECT ROWID FROM projects
                    WHERE intake_number = ?
                 ORDER BY timestamp DESC LIMIT 1)
            """, (sqlite3.Binary(content), intake_number))

        if cur.rowcount == 0:
            conn.close()
            return jsonify({"error": "Project not found"}), 404

        conn.commit()
        conn.close()
        return jsonify({"message": "OOS reference uploaded successfully"})
        #return jsonify(_make_json_safe(rows)), 200

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print("oosref_upload error:", e)
        return jsonify({"error": str(e)}), 500



@app.route('/api/fetchprojectsstatus', methods=['GET'])
@app.route('/api/fetchprojectsstatus/', methods=['GET'])  # allow trailing slash
def fetch_projects_status():
    """
    Query param:
      - automation_lead: name to match against projects.automation_qe_lead
    Returns latest project row per intake_number for that lead, plus latest
    automationdeliverables snapshot (checkboxes + ROIMonth), and key status fields.
    """
    try:
        lead = request.args.get("automation_lead", "").strip()
        if not lead:
            return jsonify({"projects": []}), 200

        conn = get_db_connection()
        cur = conn.cursor()

        # Latest project rows per intake_number for the given lead
        cur.execute("""
            SELECT p.*
              FROM projects p
              JOIN (
                    SELECT intake_number, MAX(timestamp) AS ts
                      FROM projects
                     WHERE automation_qe_lead = ?
                  GROUP BY intake_number
              ) latest
                ON latest.intake_number = p.intake_number
               AND latest.ts = p.timestamp
            ORDER BY p.intake_name COLLATE NOCASE
        """, (lead,))
        proj_rows = cur.fetchall()

        results = []
        for row in proj_rows:
            data = dict(row)
            intake_number = data.get("intake_number")

            
            # CHANGE [REUSE EXISTING API]: Pull deliverables by calling
            # the existing /api/automationdeliverables/<intake_number> endpoint
            # so we don't duplicate logic.
            deliverables = {}
            try:
                with app.test_client() as c:
                    r2 = c.get(f"/api/automationdeliverables/{intake_number}")
                    if r2.status_code == 200:
                        # That endpoint returns the latest deliverables row as JSON
                        payload = r2.get_json() or {}
                        if isinstance(payload, dict):
                            deliverables = payload
            except Exception as _e:
                # Fall back to empty deliverables on any error; no other logic changed
                deliverables = {}

            results.append({
                "intake_number": intake_number,
                "intake_name": data.get("intake_name", ""),
                "project_status": data.get("project_status", ""),
                "AutomationStatus": data.get("AutomationStatus", ""),
                "Comments": data.get("Comments", ""),
                "timestamp": data.get("timestamp", ""),
                "deliverables": {
                    "AutomationTestPlan": deliverables.get("AutomationTestPlan", ""),
                    "Checklist":          deliverables.get("Checklist", ""),
                    "ALMupdate":          deliverables.get("ALMupdate", ""),
                    "GITUpdate":          deliverables.get("GITUpdate", ""),
                    "DSR":                deliverables.get("DSR", ""),
                    "LESubmitted":        deliverables.get("LESubmitted", ""),
                    "RegressionTestLab":  deliverables.get("RegressionTestLab", ""),
                    "DashboardMetric":    deliverables.get("DashboardMetric", ""),
                    "ReleaseHandover":    deliverables.get("ReleaseHandover", ""),
                    "ProjectClosure":     deliverables.get("ProjectClosure", ""),
                    "ROIMonth":           deliverables.get("ROIMonth", ""),
                    "QETransformed":      deliverables.get("QETransformed", ""),
                    "ShiftLeft":          deliverables.get("ShiftLeft", ""),
                }
            })

        conn.close()
        return jsonify({"projects": results})
    except Exception as e:
        print("fetch_projects_status error:", e)
        return jsonify({"error": str(e)}), 500
    



@app.route('/api/projectroidownload', methods=['POST'])
def projectroi_download():
    """
    Payload:
      {
        "selections": [
          {"year": 2025, "months": ["January","February"]},
          {"year": 2024, "months": ["December"]}
        ]
      }
    Returns:
      { "rows": [ { ROI_FY, ROI_QR, ROI_month, Domain, Application, ROI,
                    AutomationFrmk, TranInitiative, FPItem,
                    intake_name, intake_number,
                    DealCount, DesignNew, DesignUpdated, ExecExecuted, Lead }, ... ] }
    """
    try:
        data = request.get_json(force=True) or {}
        selections = data.get("selections", [])
        if not isinstance(selections, list) or not selections:
            return jsonify({"error": "No selections provided"}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        out_rows = []

        for sel in selections:
            year = sel.get("year")
            months = sel.get("months") or []
            if not year or not months:
                continue

            # fetch projectroi rows for (year, months)
            rows = []
            if not months or len(months) >= 12:  # treat full-year or missing months as year-only
                sql = """
                    SELECT ROI_FY, ROI_QR, ROI_month, Domain, Application, ROI,
                           AutomationFrmk, TranInitiative, FPItem,
                          intake_name, intake_number
                      FROM projectroi
                     WHERE ROI_FY = ?
                """
                rows = cur.execute(sql, (year,)).fetchall()
            else:
                placeholders = ",".join(["?"] * len(months))
                sql = f"""
                    SELECT ROI_FY, ROI_QR, ROI_month, Domain, Application, ROI,
                          AutomationFrmk, TranInitiative, FPItem,
                           intake_name, intake_number
                      FROM projectroi
                    WHERE ROI_FY = ?
                       AND ROI_month IN ({placeholders})
                """
                rows = cur.execute(sql, [year] + months).fetchall()

            for r in rows:
                row = dict(r)

                # Build "number — name" project key for *_tccount_* tables
                proj_key = f"{row.get('intake_number','')} — {row.get('intake_name','')}"

                # Deal Count from projecttccount_TDM (latest)
                try:
                    dom = row.get("Domain")
                    frmk = row.get("AutomationFrmk")
                    appv = row.get("Application")
                    d = cur.execute(
                        """
                        SELECT AutoTDMCountExec
                        FROM projecttccount_TDM
                        WHERE project = ?
                        AND ( (Domain IS NULL AND ? IS NULL) OR Domain = ? )
                        AND ( (AutomationFrmk IS NULL AND ? IS NULL) OR AutomationFrmk = ? )
                        AND ( (application IS NULL AND ? IS NULL) OR application = ? )
                        AND ( (SavingsCategory IS NULL AND ? IS NULL) OR SavingsCategory = ? )
                        ORDER BY ROWID DESC
                        LIMIT 1
                        """,
                        (proj_key, dom, dom, frmk, frmk, appv, appv)
                    ).fetchone()
                    row["DealCount"] = (d["AutoTDMCountExec"] if isinstance(d, sqlite3.Row) else d[0]) if d else None
                except Exception:
                    row["DealCount"] = None

                # Design counts (latest)
                try:
                    dom = row.get("Domain")
                    frmk = row.get("AutomationFrmk")
                    appv = row.get("Application")
                    dn = cur.execute(
                        """
                        SELECT TotalNewAutoTCCreated, TotalExistingAutoTCUpdated
                        FROM projecttccount_Design
                        WHERE project = ?
                        AND ( (Domain IS NULL AND ? IS NULL) OR Domain = ? )
                        AND ( (AutomationFrmk IS NULL AND ? IS NULL) OR AutomationFrmk = ? )
                        AND ( (application IS NULL AND ? IS NULL) OR application = ? )
                        AND ( (SavingsCategory IS NULL AND ? IS NULL) OR SavingsCategory = ? )
                        ORDER BY ROWID DESC
                        LIMIT 1
                        """,
                        (proj_key, dom, dom, frmk, frmk, appv, appv)
                    ).fetchone()
                    if dn:
                        if isinstance(dn, sqlite3.Row):
                            row["DesignNew"] = dn.get("TotalNewAutoTCCreated")
                            row["DesignUpdated"] = dn.get("TotalExistingAutoTCUpdated")
                        else:
                            row["DesignNew"] = dn[0]
                            row["DesignUpdated"] = dn[1]
                    else:
                        row["DesignNew"] = None
                        row["DesignUpdated"] = None
                except Exception:
                    row["DesignNew"] = None
                    row["DesignUpdated"] = None

                # Execution count (LATEST) — MATCH project + Domain + AutomationFrmk + application
                try:
                    dom = row.get("Domain")
                    frmk = row.get("AutomationFrmk")
                    appv = row.get("Application")
                    ex = cur.execute(
                        """
                        SELECT TotalNewAutoTCExecuted
                        FROM projecttccount_Execution
                        WHERE project = ?
                        AND ( (Domain IS NULL AND ? IS NULL) OR Domain = ? )
                        AND ( (AutomationFrmk IS NULL AND ? IS NULL) OR AutomationFrmk = ? )
                        AND ( (application IS NULL AND ? IS NULL) OR application = ? )
                        AND ( (SavingsCategory IS NULL AND ? IS NULL) OR SavingsCategory = ? )
                        ORDER BY ROWID DESC
                        LIMIT 1
                        """,
                        (proj_key, dom, dom, frmk, frmk, appv, appv)
                    ).fetchone()
                    row["ExecExecuted"] = (ex["TotalNewAutoTCExecuted"] if isinstance(ex, sqlite3.Row) else ex[0]) if ex else None
                except Exception:
                    row["ExecExecuted"] = None

                # Lead from latest projects row by intake_number (robust join)
                try:
                    lead = cur.execute(
                        "SELECT automation_qe_lead FROM projects WHERE intake_number = ? ORDER BY timestamp DESC LIMIT 1",
                        (row.get("intake_number"),)
                    ).fetchone()
                    row["Lead"] = (lead["automation_qe_lead"] if isinstance(lead, sqlite3.Row) else lead[0]) if lead else None
                except Exception:
                    row["Lead"] = None

                out_rows.append(row)

        conn.close()
        wants_xlsx = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in (request.headers.get('Accept') or '')
        if not wants_xlsx:
            return jsonify({"rows": out_rows})

        # Build Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "ProjectROI_dump"
        headers = [
            ("FY", "ROI_FY"),
           ("ROI Reporting Qr", "ROI_QR"),
            ("ROI Reporting Month", "ROI_month"),
            ("Portfolio", "Domain"),
            ("Application", "Application"),
            ("$ Savings", "ROI"),
            ("Project", None),             # intake_name - intake_number
            ("Automation. Frmk", "AutomationFrmk"),
           ("Trans. Initiative", "TranInitiative"),
            ("FP Item?", "FPItem"),
            ("Deal Count", "DealCount"),
            ("Test Count / Calculations", None),  # "Refer Count Column"
           ("New Test Case Designed", "DesignNew"),
            ("Existing Test Case Updated", "DesignUpdated"),
            ("Test Case Execution [SIT/UAT]", "ExecExecuted"),
            ("Lead", "Lead"),
            ("Comment", None),             # blank
        ]

       # Header row with bold + orange fill
        ws.append([h[0] for h in headers])
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(fill_type="solid", fgColor="FFA500")  # orange
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        # Data rows
        for r in out_rows:
            ws.append([
                r.get("ROI_FY"),
               r.get("ROI_QR"),
                r.get("ROI_month"),
                r.get("Domain"),
                r.get("Application"),
                r.get("ROI"),
                f"{r.get('intake_number','')} — {r.get('intake_name','')}",
                r.get("AutomationFrmk"),
                r.get("TranInitiative"),
                r.get("FPItem"),
                r.get("DealCount"),
                "Refer Count Column",
                r.get("DesignNew"),
                r.get("DesignUpdated"),
               r.get("ExecExecuted"),
                r.get("Lead"),
                "",
            ])

        # Stream as XLSX
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
           download_name="ProjectROI_dump.xlsx",
        )

    except Exception as e:
        print("projectroi_download error:", e)
        return jsonify({"error": "Failed to build ROI export"}), 500    
    
@app.route('/api/harvesthours', methods=['GET'])
def api_harvesthours():
    """
    GET /api/harvesthours?siteId=...&listId=...&project=<intake_number — intake_name>
    - siteId/listId optional; if missing we use env vars in HarvestHrSP.
    - Returns: { items: [...] }
    """
    try:
        site_id = request.args.get("siteId", "").strip() or None
        list_id = request.args.get("listId", "").strip() or None
        project_key = request.args.get("project", "").strip() or None
        items = fetch_harvested_hours(site_id=site_id, list_id=list_id, project_key=project_key)
        return jsonify({"items": items}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    


@app.route('/api/projecttccount/refresh', methods=['POST'])
def refresh_projecttccount():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Introspect columns so we only write what exists (safe/surgical)
        def table_cols(name):
            c = conn.execute(f"PRAGMA table_info({name})").fetchall()
            return {row[1] for row in c}

        ptc_cols  = table_cols("projecttccount")
        tdm_cols  = table_cols("projecttccount_TDM")
        des_cols  = table_cols("projecttccount_Design")
        exe_cols  = table_cols("projecttccount_Execution")

        # CTE aggregates per spec
        query = """
        WITH keys AS (
            SELECT project, Domain, application, AutomationFrmk FROM projecttccount_TDM
            UNION
            SELECT project, Domain, application, AutomationFrmk FROM projecttccount_Design
            UNION
            SELECT project, Domain, application, AutomationFrmk FROM projecttccount_Execution
        ),
        tdm AS (
            SELECT project, Domain, application, AutomationFrmk,
                   SUM(COALESCE(AutoTDMCountExec,0)) AS AutoTDMCount
            FROM projecttccount_TDM
            GROUP BY project, Domain, application, AutomationFrmk
        ),
        des AS (
            SELECT project, Domain, application, AutomationFrmk,
                   SUM(COALESCE(TotalNewAutoTCCreated,0)) AS NewAutoTCCreated,
                   SUM(COALESCE(TotalExistingAutoTCupdated,0)) AS ExistingAutoTCUpdate,
                   SUM(COALESCE(TotalManualTCCreated,0)) AS TotalManualTCCreated,
                   MAX(DSRRef) AS DSRRefDesign
            FROM projecttccount_Design
            GROUP BY project, Domain, application, AutomationFrmk
        ),
        exe AS (
            SELECT project, Domain, application, AutomationFrmk,
                   SUM(COALESCE(ManualExecSIT_Func,0) + COALESCE(ManualExecSIT_Reg,0)) AS ManualExecSIT,
                   SUM(COALESCE(ManualExecUAT_Func,0) + COALESCE(ManualExecUAT_Reg,0)) AS ManualExecUAT,
                   SUM(COALESCE(AutoTCExecSIT_Func,0) + COALESCE(AutoTCExecSIT_Reg,0)) AS AutoTCExecutionSIT,
                   SUM(COALESCE(AutoTCExecUAT_Func,0) + COALESCE(AutoTCExecUAT_Reg,0)) AS AutoTCExecutionUAT,
                   MAX(DSRRef) AS DSRRefExec
            FROM projecttccount_Execution
            GROUP BY project, Domain, application, AutomationFrmk
        )
        SELECT
            k.project, k.Domain, k.application, k.AutomationFrmk,
            COALESCE(tdm.AutoTDMCount, 0)                   AS AutoTDMCount,
            COALESCE(des.NewAutoTCCreated, 0)               AS NewAutoTCCreated,
            COALESCE(des.ExistingAutoTCUpdate, 0)           AS ExistingAutoTCUpdate,
            COALESCE(des.TotalManualTCCreated, 0)           AS TotalManualTCCreated,
            COALESCE(exe.ManualExecSIT, 0)                  AS ManualExecSIT,
            COALESCE(exe.ManualExecUAT, 0)                  AS ManualExecUAT,
            COALESCE(exe.AutoTCExecutionSIT, 0)             AS AutoTCExecutionSIT,
            COALESCE(exe.AutoTCExecutionUAT, 0)             AS AutoTCExecutionUAT,
            COALESCE(des.DSRRefDesign, exe.DSRRefExec)      AS ChangeType,     -- spec 10
            CURRENT_TIMESTAMP                                AS LastUpdate,
            (
                SELECT p.change_type
                  FROM projects p
                 WHERE (CAST(p.intake_number AS TEXT) || '-' || COALESCE(p.intake_name,'')) = k.project
                 ORDER BY p.timestamp DESC LIMIT 1
            )                                               AS DSRRef          -- spec 9
        FROM keys k
        LEFT JOIN tdm ON tdm.project = k.project AND tdm.Domain = k.Domain AND tdm.application = k.application AND tdm.AutomationFrmk = k.AutomationFrmk
        LEFT JOIN des ON des.project = k.project AND des.Domain = k.Domain AND des.application = k.application AND des.AutomationFrmk = k.AutomationFrmk
        LEFT JOIN exe ON exe.project = k.project AND exe.Domain = k.Domain AND exe.application = k.application AND exe.AutomationFrmk = k.AutomationFrmk
        """
        rows = cur.execute(query).fetchall()
        cols = [d[0] for d in cur.description]

        # Prepare upsert (update if exists by composite key; else insert)
        key_cols = ["project","Domain","application","AutomationFrmk"]
        updated = inserted = 0

        for row in rows:
            record = dict(zip(cols, row))

            # Filter to only columns that actually exist in 'projecttccount'
            writable = {k: v for k, v in record.items() if k in ptc_cols}

            # ensure keys exist
            if not all(k in writable for k in key_cols):
                # if keys missing in schema, skip safely
                continue

            # try UPDATE
            set_cols = [c for c in writable.keys() if c not in key_cols]
            if set_cols:
                set_clause = ", ".join([f"{c}=?" for c in set_cols])
                cur.execute(
                    f"""UPDATE projecttccount
                           SET {set_clause}
                         WHERE project=? AND Domain=? AND application=? AND AutomationFrmk=?""",
                    [writable[c] for c in set_cols] + [writable["project"], writable["Domain"], writable["application"], writable["AutomationFrmk"]]
                )
                if cur.rowcount > 0:
                    updated += 1
                    continue  # done

            # if no row affected, INSERT
            insert_cols = list(writable.keys())
            placeholders = ",".join(["?"] * len(insert_cols))
            cur.execute(
                f"INSERT INTO projecttccount ({','.join(insert_cols)}) VALUES ({placeholders})",
                [writable[c] for c in insert_cols]
            )
            inserted += 1

        conn.commit()
        conn.close()
        return jsonify({"ok": True, "updated": updated, "inserted": inserted})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print("projecttccount refresh error:", e)
        return jsonify({"ok": False, "error": str(e)}), 500


# ==================== Exact-template ROI Sheet builder & download ====================

@app.route('/api/projectroisheetdownload', methods=['POST'])
def projectroisheetdownload():  # NEW
    """
    Build & download the ROI Excel using the fixed template and DB aggregates.
    Implementation delegated to ROIExcelDownload.build_roi_workbook().
    """
    try:
        payload = request.get_json(force=True, silent=False)
    except Exception:
        return jsonify({"error": "Invalid JSON"}), 400
    if not payload or "intake_number" not in payload or "intake_name" not in payload:
        return jsonify({"error": "intake_number and intake_name are required"}), 400

    # Reuse the same DB used elsewhere in this app
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "database.db")

    try:
        stream, filename = build_roi_workbook(payload, db_path)
        return send_file(
            stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print("projectroisheetdownload error:", e)
        return jsonify({"error": "Failed to build ROI workbook"}), 500




@app.route('/api/masterdashboardfetchupdate', methods=['GET', 'POST'])
def masterdashboardfetchupdate():
    """
    GET (fetch totals):
      Query params:
        mode=fetch
        scope=design|execution
        which=design_new|design_existing   (when scope=design)
        phase=DIT|SIT|UAT|Rollback|Roll Forward   (when scope=execution)
        project, Domain, AutomationFrmk, application

      Returns: { "total": <int> }

    POST (update func/reg for Design sections):
      JSON body:
        {
          "mode": "update",
          "project": "...",
          "Domain": "...",
          "AutomationFrmk": "...",
          "application": "...",
          "updates": [
            {"which":"design_new","func":<int>,"reg":<int>},
            {"which":"design_existing","func":<int>,"reg":<int>}
          ]
        }
      Returns: { "ok": True, "updated": <count> }
    """
    try:
        if request.method == 'GET':
            mode  = (request.args.get('mode') or '').lower()
            scope = (request.args.get('scope') or '').lower()
            which = (request.args.get('which') or '').strip()
            phase = (request.args.get('phase') or '').strip()

            project = (request.args.get('project') or '').strip()
            domain  = (request.args.get('Domain') or '').strip()
            frmk    = (request.args.get('AutomationFrmk') or '').strip()
            appval  = (request.args.get('application') or '').strip()

            conn = get_db_connection()
            cur = conn.cursor()

            if mode == 'fetch_manual':
                if not project:
                    return jsonify({"designed": {"func":0,"reg":0}, "executed":{"func":0,"reg":0}})
                # (a.1+a.2) first row in Design where SavingsCategory='New Test case Created'
                r1 = cur.execute("""
                    SELECT COALESCE(NewManualTC_Func,0), COALESCE(NewManualTC_Reg,0)
                      FROM projecttccount_Design
                     WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                       AND SavingsCategory='New Test case Created'
                     ORDER BY ROWID ASC LIMIT 1
                """, (project,)).fetchone()
                # (a.3+a.4) first row in Execution for this project
                r2 = cur.execute("""
                    SELECT COALESCE(ManualExecSIT_Func,0), COALESCE(ManualExecSIT_Reg,0)
                      FROM projecttccount_Execution
                     WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                     ORDER BY ROWID ASC LIMIT 1
                """, (project,)).fetchone()
                conn.close()
                return jsonify({
                  "designed": {"func": int((r1 or [0,0])[0]), "reg": int((r1 or [0,0])[1])},
                  "executed": {"func": int((r2 or [0,0])[0]), "reg": int((r2 or [0,0])[1])},
                })
            
            if mode == 'fetch_split':
                if not project:
                    return jsonify({"func": 0, "reg": 0})
                if scope == 'design':
                    # design_new → NewAutoTC_* ; design_existing → ExistingAutoTCUpdate_*
                    if which == 'design_new':
                        sql = """
                          SELECT COALESCE(NewAutoTC_Func,0), COALESCE(NewAutoTC_Reg,0)
                            FROM projecttccount_Design
                          WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                             AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                             AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                             AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                             AND SavingsCategory='New Test case Created'
                           ORDER BY ROWID DESC LIMIT 1;
                        """
                        row = cur.execute(sql, (project, domain, frmk, appval)).fetchone()
                        conn.close()
                        return jsonify({"func": int((row or [0,0])[0]), "reg": int((row or [0,0])[1])})
                    elif which == 'design_existing':
                        sql = """
                          SELECT COALESCE(ExistingAutoTCUpdate_Func,0), COALESCE(ExistingAutoTCUpdate_Reg,0)
                            FROM projecttccount_Design
                          WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                             AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                             AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                             AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                             AND SavingsCategory='Existing Test case Updated/Reused'
                           ORDER BY ROWID DESC LIMIT 1;
                        """
                        row = cur.execute(sql, (project, domain, frmk, appval)).fetchone()
                        conn.close()
                        return jsonify({"func": int((row or [0,0])[0]), "reg": int((row or [0,0])[1])})
                elif scope == 'execution':
                    # all 5 phases read: AutoTCExecUAT_Func / AutoTCExecUAT_Reg
                    # non-SOA exact match; SOA looked up via IN/LIKE when saving — here we return last row of exact base
                    base_map = {
                        "DIT": "Test Case Execution - DIT",
                        "SIT": "Test Case Execution - SIT",
                        "UAT": "Test Case Execution - UAT",
                        "Rollback": "Test Case Execution - RollBack",
                        "Roll Forward": "Test Case Execution - RollForward - SIT",
                    }
                    sc = base_map.get(phase)
                    if not sc:
                        conn.close(); return jsonify({"func": 0, "reg": 0})
                    # Prefer exact SavingsCategory for consistency with non-SOA; ok if 0 rows (returns 0/0)
                    row = cur.execute("""
                        SELECT COALESCE(AutoTCExecUAT_Func,0), COALESCE(AutoTCExecUAT_Reg,0)
                          FROM projecttccount_Execution
                         WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                           AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                           AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                           AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                           AND SavingsCategory = ?
                         ORDER BY ROWID DESC LIMIT 1;
                    """, (project, domain, frmk, appval, sc)).fetchone()
                    conn.close()
                    return jsonify({"func": int((row or [0,0])[0]), "reg": int((row or [0,0])[1])})
                

            if mode != 'fetch' or not project or not domain or not frmk:
                return jsonify({"total": 0})

            is_soa = (frmk.strip().lower() == 'soa api')



            def _single_val(sql, params):
                row = cur.execute(sql, params).fetchone()
                if not row:
                    return 0
                # support row as Row or tuple
                val = row[0] if not isinstance(row, sqlite3.Row) else list(row)[0]
                try:
                    return int(val or 0)
                except Exception:
                    return 0

            total = 0

            # =========================
            # NEW: Manual tab fetch (designed + executed) — same selection used when saving
            # =========================




            if scope == 'design':
                # projecttccount_Design
                if which == 'design_new':
                    if not is_soa:
                        sql = """
                          SELECT COALESCE(TotalNewAutoTCCreated,0)
                            FROM projecttccount_Design
                           WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                            AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                            AND SavingsCategory='New Test case Created'
                           ORDER BY ROWID DESC LIMIT 1;
                        """
                        total = _single_val(sql, (project, domain, frmk, appval))
                    else:
                        sql = """
                          SELECT COALESCE(SUM(TotalNewAutoTCCreated),0)
                            FROM projecttccount_Design
                           WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                            AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                             AND SavingsCategory IN (
                               'New Test case Created (New SOA Package)',
                               'New Test case Created (Existing SOA Package update)'
                             );
                        """
                        total = _single_val(sql, (project, domain, frmk, appval))

                elif which == 'design_existing':
                    if not is_soa:
                        sql = """
                          SELECT COALESCE(TotalExistingAutoTCUpdated,0)
                            FROM projecttccount_Design
                           WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                             AND SavingsCategory='Existing Test case Updated/Reused'
                           ORDER BY ROWID DESC LIMIT 1;
                        """
                        total = _single_val(sql, (project, domain, frmk, appval))
                    else:
                        sql = """
                          SELECT COALESCE(SUM(TotalExistingAutoTCUpdated),0)
                            FROM projecttccount_Design
                             WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                             AND SavingsCategory IN (
                               'Existing Test case Updated/Reused (New SOA Package)',
                               'Existing Test case Updated/Reused (Existing SOA Package update)'
                             );
                        """
                        total = _single_val(sql, (project, domain, frmk, appval))

            elif scope == 'execution':
                # projecttccount_Execution
                # Map UI phase -> SavingsCategory (non-SOA) or IN (...) (SOA)
                def _phase_clause(ph):
                    if ph == 'DIT':
                        return ("Test Case Execution - DIT",
                                ["Test Case Execution - DIT (Existing SOA Package update)",
                                 "Test Case Execution - DIT (New SOA Package)"])
                    if ph == 'SIT':
                        return ("Test Case Execution - SIT",
                                ["Test Case Execution - SIT (Existing SOA Package update)",
                                 "Test Case Execution - SIT (New SOA Package)"])
                    if ph == 'UAT':
                        return ("Test Case Execution - UAT",
                                ["Test Case Execution - UAT (Existing SOA Package update)",
                                 "Test Case Execution - UAT (New SOA Package)"])
                    if ph == 'Rollback':
                        return ("Test Case Execution - RollBack",
                                ["Test Case Execution - RollBack (Existing SOA Package update)",
                                 "Test Case Execution - RollBack (New SOA Package)"])
                    if ph == 'Roll Forward':
                        return ("Test Case Execution - RollForward",
                                ["Test Case Execution - RollForward (Existing SOA Package update)",
                                 "Test Case Execution - RollForward (New SOA Package)"])
                    return (None, None)

                cat, soa_list = _phase_clause(phase)
                if cat is None:
                    conn.close()
                    return jsonify({"total": 0})

                if not is_soa:
                    sql = """
                      SELECT COALESCE(SUM(TotalNewAutoTCExecuted),0)
                        FROM projecttccount_Execution
                                                   WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                         AND SavingsCategory = ?;
                    """
                    total = _single_val(sql, (project, domain, frmk, appval, cat))
                else:
                    placeholders = ",".join(["?"] * len(soa_list))
                    sql = f"""
                      SELECT COALESCE(SUM(TotalNewAutoTCExecuted),0)
                        FROM projecttccount_Execution
                                                   WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                         AND SavingsCategory IN ({placeholders});
                    """
                    total = _single_val(sql, (project, domain, frmk, appval, *soa_list))

            conn.close()
            return jsonify({"total": int(total)})

        # ------------------------ POST: update ------------------------
        payload = request.get_json(force=True) or {}

        # [NEW MANUAL BRANCH] handle Manual-tab updates (a.1–a.4)
        if (payload.get("mode") or "").lower() == "update_manual":
            project = (payload.get('project') or '').strip()
            if not project:
                return jsonify({"error": "Invalid payload"}), 400

            updated = 0
            with _db_write_lock:
                conn = get_db_connection()
                cur = conn.cursor()
                try:
                    # a.1 + a.2: projecttccount_Design, first row where SavingsCategory='New Test case Created'
                    cur.execute("""
                        SELECT ROWID FROM projecttccount_Design
                         WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                           AND SavingsCategory='New Test case Created'
                         ORDER BY ROWID ASC LIMIT 1
                    """, (project,))
                    r = cur.fetchone()
                    if r:
                        rid = r[0] if not isinstance(r, sqlite3.Row) else list(r)[0]
                        vfunc = int((payload.get("updates") or {}).get("design_new_manual", {}).get("func") or 0)
                        vreg  = int((payload.get("updates") or {}).get("design_new_manual", {}).get("reg")  or 0)
                        cur.execute("""
                            UPDATE projecttccount_Design
                               SET NewManualTC_Func=?, NewManualTC_Reg=?
                             WHERE ROWID=?
                        """, (vfunc, vreg, rid))
                        updated += cur.rowcount

                    # a.3 + a.4: projecttccount_Execution, first row for this project
                    cur.execute("""
                        SELECT ROWID FROM projecttccount_Execution
                         WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                         ORDER BY ROWID ASC LIMIT 1
                    """, (project,))
                    r = cur.fetchone()
                    if r:
                        rid = r[0] if not isinstance(r, sqlite3.Row) else list(r)[0]
                        vfunc = int((payload.get("updates") or {}).get("executed_manual", {}).get("func") or 0)
                        vreg  = int((payload.get("updates") or {}).get("executed_manual", {}).get("reg")  or 0)
                        cur.execute("""
                            UPDATE projecttccount_Execution
                               SET ManualExecSIT_Func=?, ManualExecSIT_Reg=?
                             WHERE ROWID=?
                        """, (vfunc, vreg, rid))
                        updated += cur.rowcount

                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise
                finally:
                    conn.close()
            return jsonify({"ok": True, "updated": int(updated)})

        # -------- original DESIGN update branch (unchanged) --------
        if (payload.get("mode") or "").lower() != "update":
            return jsonify({"error": "Unsupported mode"}), 400

        project = (payload.get('project') or '').strip()
        domain  = (payload.get('Domain') or '').strip()
        frmk    = (payload.get('AutomationFrmk') or '').strip()
        appval  = (payload.get('application') or '').strip()
        updates = payload.get('updates', [])

        if not project or not domain or not frmk or not isinstance(updates, list):
            return jsonify({"error": "Invalid payload"}), 400

        is_soa = (frmk.strip().lower() == 'soa api')
        updated = 0

        with _db_write_lock:
            conn = get_db_connection()
            cur = conn.cursor()

            def _update_design(which, func_col, reg_col, base_cat, soa_new_cat, soa_exist_cat, func_val, reg_val):
                # If not SOA → single category; if SOA → try 'New' row first, else update 'Existing' row
                if not is_soa:
                    sql = f"""
                      UPDATE projecttccount_Design
                         SET {func_col} = ?, {reg_col} = ?
                       WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                        AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                        AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                        AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                         AND SavingsCategory=?
                    """
                    cur.execute(sql, (int(func_val), int(reg_val), project, domain, frmk, appval, base_cat))
                    return cur.rowcount
                else:
                    sql_new = f"""
                      UPDATE projecttccount_Design
                         SET {func_col} = ?, {reg_col} = ?
                       WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                        AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                        AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                        AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                         AND SavingsCategory=?
                    """
                    cur.execute(sql_new, (int(func_val), int(reg_val), project, domain, frmk, appval, soa_new_cat))
                    if cur.rowcount == 0:
                        sql_old = f"""
                          UPDATE projecttccount_Design
                             SET {func_col} = ?, {reg_col} = ?
                           WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                            AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                             AND SavingsCategory=?
                        """
                        cur.execute(sql_old, (int(func_val), int(reg_val), project, domain, frmk, appval, soa_exist_cat))
                    return cur.rowcount

            for u in updates:
                which = (u.get("which") or "").strip()
                func  = int(u.get("func") or 0)
                reg   = int(u.get("reg") or 0)

                if which == "design_new":
                    updated += _update_design(
                        which,
                        "NewAutoTC_Func", "NewAutoTC_Reg",
                        "New Test case Created",
                        "New Test case Created (New SOA Package)",
                        "New Test case Created (Existing SOA Package update)",
                        func, reg
                    )
                elif which == "design_existing":
                    updated += _update_design(
                        which,
                        "ExistingAutoTCUpdate_Func", "ExistingAutoTCUpdate_Reg",
                        "Existing Test case Updated/Reused",
                        "Existing Test case Updated/Reused (New SOA Package)",
                        "Existing Test case Updated/Reused (Existing SOA Package update)",
                        func, reg
                    )

            # [NEW EXECUTION UPDATES] handle b.2 phases for projecttccount_Execution
            exec_updates = payload.get('exec_updates', [])
            if isinstance(exec_updates, list) and exec_updates and project and domain and frmk and appval:
                # Map UI phase -> SavingsCategory (write-time).
                # Per instructions, UAT maps to SIT row on update.
                def _phase_sc(ph):
                    ph = (ph or "").strip()
                    if ph == "DIT":
                        return "Test Case Execution - DIT"
                    if ph == "SIT":
                        return "Test Case Execution - SIT"
                    if ph == "UAT":
                        return "Test Case Execution - UAT"        # per spec
                    if ph == "Rollback":
                        return "Test Case Execution - RollBack"
                    if ph == "Roll Forward":
                        return "Test Case Execution - RollForward"
                    return None

                for e in exec_updates:
                    phase = (e.get("phase") or "").strip()
                    func  = int(e.get("func") or 0)
                    reg   = int(e.get("reg") or 0)
                    sc    = _phase_sc(phase)
                    if not sc:
                        continue

                    if is_soa and phase == "DIT":
                        # SOA: SavingsCategory includes "(New SOA Package)" etc. → use LIKE
                        cur.execute("""
                            UPDATE projecttccount_Execution
                               SET AutoTCExecUAT_Func=?, AutoTCExecUAT_Reg=?
                             WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                               AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                               AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                               AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                               AND UPPER(TRIM(SavingsCategory)) LIKE UPPER(TRIM(?))
                        """, (func, reg, project, domain, frmk, appval, sc + " (%SOA%)"))
                        if cur.rowcount == 0:
                            # fallback: any SOA variant
                            cur.execute("""
                                UPDATE projecttccount_Execution
                                   SET AutoTCExecUAT_Func=?, AutoTCExecUAT_Reg=?
                                 WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                   AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                   AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                   AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                                   AND UPPER(TRIM(SavingsCategory)) LIKE UPPER(TRIM(?))
                            """, (func, reg, project, domain, frmk, appval, "%SOA%"))
                    else:
                        cur.execute("""
                            UPDATE projecttccount_Execution
                               SET AutoTCExecUAT_Func=?, AutoTCExecUAT_Reg=?
                             WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                               AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                               AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                               AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                               AND SavingsCategory = ?
                        """, (func, reg, project, domain, frmk, appval, sc))
                    updated += cur.rowcount

            conn.commit()
            conn.close()
        return jsonify({"ok": True, "updated": int(updated)})

    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        print("masterdashboardfetchupdate error:", e)
        return jsonify({"error": "Failed"}), 500


'''
@app.route('/api/masterdashboardfetchupdate', methods=['GET', 'POST'])
def masterdashboardfetchupdate():
    """
    GET (fetch totals):
      Query params:
        mode=fetch
        scope=design|execution
        which=design_new|design_existing   (when scope=design)
        phase=DIT|SIT|UAT|Rollback|Roll Forward   (when scope=execution)
        project, Domain, AutomationFrmk, application

      Returns: { "total": <int> }

    POST (update func/reg for Design sections):
      JSON body:
        {
          "mode": "update",
          "project": "...",
          "Domain": "...",
          "AutomationFrmk": "...",
          "application": "...",
          "updates": [
            {"which":"design_new","func":<int>,"reg":<int>},
            {"which":"design_existing","func":<int>,"reg":<int>}
          ]
        }
      Returns: { "ok": True, "updated": <count> }
    """
    try:
        if request.method == 'GET':
            mode  = (request.args.get('mode') or '').lower()
            scope = (request.args.get('scope') or '').lower()
            which = (request.args.get('which') or '').strip()
            phase = (request.args.get('phase') or '').strip()

            project = (request.args.get('project') or '').strip()
            domain  = (request.args.get('Domain') or '').strip()
            frmk    = (request.args.get('AutomationFrmk') or '').strip()
            appval  = (request.args.get('application') or '').strip()

            if mode != 'fetch' or not project or not domain or not frmk:
                return jsonify({"total": 0})

            is_soa = (frmk.strip().lower() == 'soa api')

            conn = get_db_connection()
            cur = conn.cursor()

            def _single_val(sql, params):
                row = cur.execute(sql, params).fetchone()
                if not row:
                    return 0
                # support row as Row or tuple
                val = row[0] if not isinstance(row, sqlite3.Row) else list(row)[0]
                try:
                    return int(val or 0)
                except Exception:
                    return 0

            total = 0

            if scope == 'design':
                # projecttccount_Design
                if which == 'design_new':
                    if not is_soa:
                        sql = """
                          SELECT COALESCE(TotalNewAutoTCCreated,0)
                            FROM projecttccount_Design
                           WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                            AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                            AND SavingsCategory='New Test case Created'
                           ORDER BY ROWID DESC LIMIT 1;
                        """
                        total = _single_val(sql, (project, domain, frmk, appval))
                    else:
                        sql = """
                          SELECT COALESCE(SUM(TotalNewAutoTCCreated),0)
                            FROM projecttccount_Design
                           WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                            AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                             AND SavingsCategory IN (
                               'New Test case Created (New SOA Package)',
                               'New Test case Created (Existing SOA Package update)'
                             );
                        """
                        total = _single_val(sql, (project, domain, frmk, appval))

                elif which == 'design_existing':
                    if not is_soa:
                        sql = """
                          SELECT COALESCE(TotalExistingAutoTCUpdated,0)
                            FROM projecttccount_Design
                           WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                             AND SavingsCategory='Existing Test case Updated/Reused'
                           ORDER BY ROWID DESC LIMIT 1;
                        """
                        total = _single_val(sql, (project, domain, frmk, appval))
                    else:
                        sql = """
                          SELECT COALESCE(SUM(TotalExistingAutoTCUpdated),0)
                            FROM projecttccount_Design
                             WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                             AND SavingsCategory IN (
                               'Existing Test case Updated/Reused (New SOA Package)',
                               'Existing Test case Updated/Reused (Existing SOA Package update)'
                             );
                        """
                        total = _single_val(sql, (project, domain, frmk, appval))

            elif scope == 'execution':
                # projecttccount_Execution
                # Map UI phase -> SavingsCategory (non-SOA) or IN (...) (SOA)
                def _phase_clause(ph):
                    if ph == 'DIT':
                        return ("Test Case Execution - DIT",
                                ["Test Case Execution - DIT (Existing SOA Package update)",
                                 "Test Case Execution - DIT (New SOA Package)"])
                    if ph == 'SIT':
                        return ("Test Case Execution - SIT",
                                ["Test Case Execution - SIT (Existing SOA Package update)",
                                 "Test Case Execution - SIT (New SOA Package)"])
                    if ph == 'UAT':
                        return ("Test Case Execution - UAT",
                                ["Test Case Execution - UAT (Existing SOA Package update)",
                                 "Test Case Execution - UAT (New SOA Package)"])
                    if ph == 'Rollback':
                        return ("Test Case Execution - RollBack",
                                ["Test Case Execution - RollBack (Existing SOA Package update)",
                                 "Test Case Execution - RollBack (New SOA Package)"])
                    if ph == 'Roll Forward':
                        return ("Test Case Execution - RollForward",
                                ["Test Case Execution - RollForward (Existing SOA Package update)",
                                 "Test Case Execution - RollForward (New SOA Package)"])
                    return (None, None)

                cat, soa_list = _phase_clause(phase)
                if cat is None:
                    conn.close()
                    return jsonify({"total": 0})

                if not is_soa:
                    sql = """
                      SELECT COALESCE(SUM(TotalNewAutoTCExecuted),0)
                        FROM projecttccount_Execution
                                                   WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                         AND SavingsCategory = ?;
                    """
                    total = _single_val(sql, (project, domain, frmk, appval, cat))
                else:
                    placeholders = ",".join(["?"] * len(soa_list))
                    sql = f"""
                      SELECT COALESCE(SUM(TotalNewAutoTCExecuted),0)
                        FROM projecttccount_Execution
                                                   WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                                AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                                AND UPPER(TRIM(application)) = UPPER(TRIM(?))
                         AND SavingsCategory IN ({placeholders});
                    """
                    total = _single_val(sql, (project, domain, frmk, appval, *soa_list))

            conn.close()
            return jsonify({"total": int(total)})

        # ------------------------ POST: update design func/reg ------------------------
        payload = request.get_json(force=True) or {}
        if (payload.get("mode") or "").lower() != "update":
            return jsonify({"error": "Unsupported mode"}), 400

        project = (payload.get('project') or '').strip()
        domain  = (payload.get('Domain') or '').strip()
        frmk    = (payload.get('AutomationFrmk') or '').strip()
        appval  = (payload.get('application') or '').strip()
        updates = payload.get('updates', [])



        if not project or not domain or not frmk or not isinstance(updates, list):
            return jsonify({"error": "Invalid payload"}), 400

        is_soa = (frmk.strip().lower() == 'soa api')
        updated = 0

        with _db_write_lock:
            conn = get_db_connection()
            cur = conn.cursor()


            def _update_design(which, func_col, reg_col, base_cat, soa_new_cat, soa_exist_cat, func_val, reg_val):
                # If not SOA → single category; if SOA → try 'New' row first, else update 'Existing' row
                if not is_soa:
                    sql = f"""
                      UPDATE projecttccount_Design
                         SET {func_col} = ?, {reg_col} = ?
                       WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                        AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                        AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                        AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                         AND SavingsCategory=?
                    """
                    cur.execute(sql, (int(func_val), int(reg_val), project, domain, frmk, appval, base_cat))
                    return cur.rowcount
                else:
                    sql_new = f"""
                      UPDATE projecttccount_Design
                         SET {func_col} = ?, {reg_col} = ?
                       WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                        AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                        AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                        AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                         AND SavingsCategory=?
                    """
                    cur.execute(sql_new, (int(func_val), int(reg_val), project, domain, frmk, appval, soa_new_cat))
                    if cur.rowcount == 0:
                        sql_old = f"""
                          UPDATE projecttccount_Design
                             SET {func_col} = ?, {reg_col} = ?
                           WHERE UPPER(TRIM(project)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(Domain))  = UPPER(TRIM(?))
                            AND UPPER(TRIM(AutomationFrmk)) = UPPER(TRIM(?))
                            AND UPPER(TRIM(application))    = UPPER(TRIM(?))
                             AND SavingsCategory=?
                        """
                        cur.execute(sql_old, (int(func_val), int(reg_val), project, domain, frmk, appval, soa_exist_cat))
                    return cur.rowcount

            for u in updates:
                which = (u.get("which") or "").strip()
                func  = int(u.get("func") or 0)
                reg   = int(u.get("reg") or 0)

                if which == "design_new":
                    updated += _update_design(
                        which,
                        "NewAutoTC_Func", "NewAutoTC_Reg",
                        "New Test case Created",
                        "New Test case Created (New SOA Package)",
                        "New Test case Created (Existing SOA Package update)",
                        func, reg
                    )
                elif which == "design_existing":
                    updated += _update_design(
                        which,
                        "ExistingAutoTCUpdate_Func", "ExistingAutoTCUpdate_Reg",
                        "Existing Test case Updated/Reused",
                        "Existing Test case Updated/Reused (New SOA Package)",
                        "Existing Test case Updated/Reused (Existing SOA Package update)",
                        func, reg
                    )

            conn.commit()
            conn.close()
        return jsonify({"ok": True, "updated": int(updated)})

    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        print("masterdashboardfetchupdate error:", e)
        return jsonify({"error": "Failed"}), 500
'''

    
# ---------------------------------------------------------------------
# NEW: Automation Metrics data fetch (Creation/Execution)
# ---------------------------------------------------------------------
@app.route('/api/automationmetricsdatafetch', methods=['GET'])
def automationmetricsdatafetch():
    """
    Query params:
      - tab: 'creation' | 'execution'
      - view: 'consolidated' | 'monthly'
      - fy: (optional) e.g. 'FY-2024' or 'All Years' or 'FY-2026 (To-Date)'
      - breakdown: (optional) UI selection string; reserved for future grouping
    Returns:
      { headers: [..], rows: [[..],[..], ...] }
    """
    try:
        tab = (request.args.get("tab") or "").strip().lower()
        view = (request.args.get("view") or "").strip().lower()
        fy = (request.args.get("fy") or "").strip()
        breakdown = (request.args.get("breakdown") or "").strip()

        if tab not in ("creation", "execution"):
            return jsonify({"error": "invalid tab"}), 400
        if view not in ("consolidated", "monthly"):
            return jsonify({"error": "invalid view"}), 400

        if view == "consolidated":
            # sheets: Creation / Execution
            payload = get_sheet_table_consolidated(tab)
        else:
            # sheets: MonthlyCreation / MonthlyExecution
            payload = get_sheet_table_monthly(tab, fy=fy, breakdown=breakdown)

        return jsonify(payload), 200
    except Exception as e:
        print("automationmetricsdatafetch error:", e)
        return jsonify({"headers": [], "rows": []}), 200
    
@app.route('/api/fetchPendingEstimates', methods=['GET'])
def fetch_pending_estimates():
    """
    Returns rows from `projects` where AutoEstimateStatus = 'Assigned For Estimation'.
    Columns:
      - intake_number, intake_name  -> used to render "<intake_number> - <intake_name>"
      - automation_qe_lead          -> column 2
      - AutoEstimateDueDate         -> column 3
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        rows = cur.execute(
            """
            SELECT intake_number, intake_name, automation_qe_lead, AutoEstimateDueDate
              FROM projects
             WHERE AutoEstimateStatus = 'Assigned For Estimation'
          ORDER BY intake_name COLLATE NOCASE
            """
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        print("fetch_pending_estimates error:", e)
        return jsonify([]), 200


@app.route('/api/fetchPendingLeadAssignment', methods=['GET'])
def fetch_pending_lead_assignment():
    """
    Returns projects where:
      - AutoEstimateStatus = 'Sent'
      - automation_qe_lead is NULL/blank
      - (optional) release filter matching ?release=<name>
    Columns:
      - intake_number, intake_name (from projects)
      - domain (from projects)
      - release (from projects)
      - sit1date (from projects_timeline for same intake_number; latest row if multiple)
    """
    try:
        release_filter = (request.args.get("release") or "").strip()

        conn = get_db_connection()
        cur = conn.cursor()

        # LEFT JOIN timeline to fetch the latest sit1date per intake_number if table exists
        # (safe if projects_timeline not present)
        try:
            cur.execute("PRAGMA table_info(projects_timeline)")
            timeline_exists = bool(cur.fetchall())
        except Exception:
            timeline_exists = False

        if timeline_exists:
            base_sql = f"""
                SELECT p.intake_number,
                       p.intake_name,
                       p.domain,
                       p.release,
                       (
                          SELECT pt.sit1_date
                            FROM projects_timeline pt
                           WHERE pt.intake_number = p.intake_number
                        ORDER BY ROWID DESC
                           LIMIT 1
                       ) AS sit1_date
                  FROM projects p
                 WHERE p.AutoEstimateStatus = 'Sent'
                   AND (p.automation_qe_lead IS NULL OR TRIM(p.automation_qe_lead) = '')
            """
        else:
            base_sql = """
                SELECT p.intake_number,
                       p.intake_name,
                       p.domain,
                       p.release,
                       NULL AS sit1_date
                  FROM projects p
                 WHERE p.AutoEstimateStatus = 'Sent'
                   AND (p.automation_qe_lead IS NULL OR TRIM(p.automation_qe_lead) = '')
            """

        params = []
        if release_filter:
            base_sql += " AND p.release = ?"
            params.append(release_filter)

        base_sql += " GROUP BY p.intake_number, p.intake_name, p.domain, p.release ORDER BY p.intake_name COLLATE NOCASE"

        rows = cur.execute(base_sql, params).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        print("fetch_pending_lead_assignment error:", e)
        return jsonify([]), 200


@app.route('/api/countLeadActiveAssignments', methods=['GET'])
def count_lead_active_assignments():
    """
    Query params:
      - name   : resource display name to match projects.automation_qe_lead
      - intake : intake_number (as shown in col 1 on the UI row)
    Counts rows in `projects` where:
      automation_qe_lead = :name AND intake_number = :intake
      AND AutomationStatus is one of the specified active statuses.
    Returns: { "count": <int> }
    """
    try:
        name   = (request.args.get("name") or "").strip()
        #intake = (request.args.get("intake") or "").strip()
        if not name:
            return jsonify({"count": 0})
        # Build IN-list safely
        active_statuses = [
            "SRD/BRD Phase",
            "Automation Test Planning",
            "Test Design - ConformIQ",
            "Scripting & Dry Run - Robot / Selenium / UFT / SOA / Scriptless",
            "Test Execution  - DIT",
            "Test Execution  - SIT",
            "Test Execution  - UAT",
        ]
        placeholders = ",".join(["?"] * len(active_statuses))

        conn = get_db_connection()
        cur = conn.cursor()

        #if intake:
        sql = f"""
                SELECT COUNT(1)
                  FROM projects
                 WHERE automation_qe_lead = ?
                 AND COALESCE(AutomationStatus,'') IN ({placeholders})
        """
        params = [name] + active_statuses
        #else:
        row = cur.execute(sql, params).fetchone()
        conn.close()

        count_val = (row[0] if row and not isinstance(row, sqlite3.Row) else (row["COUNT(1)"] if row else 0)) or 0
        return jsonify({"count": int(count_val)})
    except Exception as e:
        print("count_lead_active_assignments error:", e)
        return jsonify({"count": 0}), 200


@app.route('/api/run_automation_sql', methods=['POST'])
def run_automation_sql():
    """
    Body: { "query_key": "<name>" }
    Reads AutomationQueries.sql and executes the named section.
    For now, if no markers exist and query_key == 'refresh_master_inventory',
    the entire file is executed (back-compat).
    """
    try:
        payload = request.get_json(force=True) or {}
        key = (payload.get("query_key") or "").strip()
        if not key:
            return jsonify({"ok": False, "error": "query_key is required"}), 400

        sql_text = _read_sql_section(AUTOMATION_SQL_FILE, key)

        with _db_write_lock:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.executescript(sql_text)   # executes multi-statement SQL safely
            conn.commit()
            conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        print("run_automation_sql error:", e)
        return jsonify({"ok": False, "error": str(e)}), 500
    

@app.route("/api/powerbi/masterinv_filters", methods=["GET"])
def powerbi_masterinv_filters():
    """
    Optional query params:
      Domain, Application, AutomationFrmk  (single values for basic filters)
    Returns: { "filters": [ ... powerbi basic filter JSON ... ] }
    """
    table = "masterautoinventory"
    filter_map = {
        "Domain": "Domain",
        "Application": "Application",
        "AutomationFrmk": "AutomationFrmk",
    }
    out = []
    for qp, col in filter_map.items():
        val = (request.args.get(qp) or "").strip()
        if not val:
            continue
        out.append({
            "$schema": "http://powerbi.com/product/schema#basic",
            "target": {"table": table, "column": col},
            "operator": "In",
            "values": [val],
            "filterType": 1,
            "requireSingleSelection": True
        })
    return jsonify({"filters": out})


@app.route("/api/powerbi/masterinv_embed", methods=["GET"])
def powerbi_masterinv_embed():
    """
    Returns:
      { "embed_url": "<publish-to-web or embed URL>", "filters": [...] }
    Embed URL is read from env POWERBI_MASTERINV_PTW_URL for flexibility.
    """
    # Reuse the filters endpoint internally to avoid duplicating logic
    with app.test_request_context():
        filters_resp = powerbi_masterinv_filters()
    try:
        filters_json = filters_resp.json if hasattr(filters_resp, "json") else filters_resp.get_json()
    except Exception:
        filters_json = {"filters": []}

    embed_url = os.environ.get("POWERBI_MASTERINV_PTW_URL", "").strip()
    if not embed_url:
        # Safe placeholder; UI will still render the area and you can set the env later
        return jsonify({
            "embed_url": "",
            "filters": filters_json.get("filters", []),
            "warning": "Set env var POWERBI_MASTERINV_PTW_URL to your publish-to-web or embed URL"
        })

    return jsonify({
        "embed_url": embed_url,
        "filters": filters_json.get("filters", [])
    })


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------
if __name__ == '__main__':
    app.run(debug=False, threaded=True)
