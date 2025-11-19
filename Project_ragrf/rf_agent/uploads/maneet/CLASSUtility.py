
import time
from datetime import datetime
from threading import Thread

import openpyxl
import traceback

from robot.libraries.BuiltIn import BuiltIn

CLASS_TDM_Deals = []


def custom_extract_excel_data(excel_path,sheet_name,tc_name,load_type=1):
    """

    Load the excel sheet and filters the "TestCaseName" for the desired value and returns a listed dictionary rows

    Parameters:
    excelPath (str): Location of Workbook on filesystem
    tc_name (str): testcase name value to search for
    sheetName (str): the name of the Worksheet to get the values from

    Returns:

    excel_dict (dict)
    """
    try:
        wb_cust_obj=openpyxl.load_workbook(excel_path)
        print(excel_path)
        print(sheet_name)
        sheet_obj=wb_cust_obj[sheet_name]
        excel_dict = {}
        match_flag = 0
        heading_row_num = 1
        if load_type == 1:
            for row_itr in range(1,sheet_obj.max_row+1):
                cell_itr = sheet_obj.cell(row=row_itr,column=1).value
                if tc_name == cell_itr:
                    for col_itr in range(1,sheet_obj.max_column+1):
                        key = sheet_obj.cell(row=heading_row_num,column=col_itr).value
                        value = sheet_obj.cell(row=row_itr,column=col_itr).value
                        if value == None:
                            excel_dict[key] = "[Dont care]"
                        else:
                            excel_dict[key] = str(value)
                    break
        if load_type == 2:
            for row_itr in range(1,sheet_obj.max_row+1):
                cell_itr = sheet_obj.cell(row=row_itr,column=1).value
                if tc_name == cell_itr:
                    for col_itr in range(1,sheet_obj.max_column+1):
                        key = sheet_obj.cell(row=heading_row_num,column=col_itr).value
                        if match_flag == 0:
                            excel_dict[key] = []
                        value = sheet_obj.cell(row=row_itr,column=col_itr).value
                        if value == None:
                            excel_dict[key].append("[Dont care]")
                        else:
                            excel_dict[key].append(str(value))
                    match_flag = 1
        return excel_dict
    except:
        import sys
        traceback.print_exception(*sys.exc_info())   
        return {} 

def keep_coins_connection_alive(ping_time='2s'):
    interval_seconds = int(ping_time[:-1])
    keep_alive_thread = Thread(target=_keep_alive, args=(interval_seconds,))
    keep_alive_thread.daemon = True
    keep_alive_thread.start()

def _keep_alive(interval):
    selenium_lib = BuiltIn().get_library_instance('SeleniumLibrary')
    print(f"Session id: {selenium_lib.driver.session_id}")
    for _ in range(1000):
        try:
            # dummy ping
            print(f"ping at  {datetime.now().strftime('%H:%M:%S')} {selenium_lib.driver.current_url}")
            selenium_lib.execute_javascript("return 1;")
            time.sleep(interval)
        except Exception as e:
            print(f"Keep coins alive failed: {e}")
            break
        
def fn_validate_xpath(xpath):
    """
        Check if given xpath exist in the page.
        Input: xpath ->
        Return status pass or Fail
    """

    try:
        selenium_library = get_driver()
        selenium_library.driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


def fn_retrieve_value_from_xpath(xpath):
    """
        This function is for retrieving the values from the XPath
    """
    status = fn_validate_xpath(xpath)
    value = "None"

    if status != False:
        try:
            selenium_library = get_driver()
            value = selenium_library.driver.find_element_by_xpath(xpath).get_attribute("value")
            print(value)
            if value == None:
                value = selenium_library.driver.find_element_by_xpath(xpath).text

        except:
            traceback.print_exception(*sys.exc_info())
            return str(value)
        return str(value)
    else:
        value = "Does not exist"
        return str(value)