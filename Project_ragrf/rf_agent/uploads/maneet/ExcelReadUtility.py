import win32com.client
import time

from openpyxl import load_workbook



def Write_Value_to_Excel(file_path, sheetname, cellName, cellvalue):
    """
            This Function writes values to given excel w.r.t Sheet, row, column
            openpyxl - writes calculation result
    """
    wb = load_workbook(file_path)
    ws = wb[sheetname]
    ws[cellName] = cellvalue
    wb.save(file_path)

def Read_Final_Result_Excel(file_path, sheetname, resultCellName):
    """
         This Function reads values in given excel w.r.t Sheet, row, column
           openpyxl - reads calculation result
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheetname]
    result = ws[resultCellName].value
    return result

def Open_Excel_In_Win32Com(file_path):
    """
            This Function opens Excel using Win32Com
            It refreshes the calculation result, saved using openpyxl
            so that if fetches the value instead of formula while reading cell value

    """
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.open(file_path)
    wb.RefreshAll()
    time.sleep(3)
    wb.Save()
    wb.Close(False)
    excel.Quit()



