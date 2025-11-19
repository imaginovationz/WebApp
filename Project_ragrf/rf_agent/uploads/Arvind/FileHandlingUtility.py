from robot.libraries.BuiltIn import BuiltIn
import openpyxl
import  os
from openpyxl import Workbook
from openpyxl import load_workbook
# from win32comext.propsys.pscon import PKEY_IsFlagged
import pandas as pd
import numpy as np
from numpy import str
import re
import traceback
import sys
import json
import string





def getTestDataFromSheetPd(excelPath, testCaseID, worksheetName):
    """
    
    Load the excel sheet and filters the "TestCaseName" for the desired value and returns a listed dictionary rows
    
    Parameters:
    excelPath (str): Location of Workbook on filesystem
    testCaseID (str): testcase Id value to search for
    worksheetName (str): the name of the Worksheet to get the values from
    
    Returns:
    
    mtg_sheet_filtered_dict (dict): dictionary of the values in the found cells - 
                                    key (str): column name
                                    values (list): cell values
    
    """
    mtg_sheet = pd.read_excel(excelPath, worksheetName, dtype=str)
    mtg_sheet_filtered = mtg_sheet[mtg_sheet['TestCaseName'].eq(testCaseID)]
    mtg_sheet_filtered = mtg_sheet_filtered.replace([np.nan], [None])
    mtg_sheet_filtered_dict = mtg_sheet_filtered.to_dict('list')
    print("Dict Without Customer Name - Pandas")
    print(mtg_sheet_filtered_dict)
    return mtg_sheet_filtered_dict

def custom_extract_excel_data_for_single_rows(excel_path,sheet_name,tc_name):
    try:
        wb_cust_obj=openpyxl.load_workbook(excel_path)
        sheet_obj=wb_cust_obj[sheet_name]
        excel_dict = {}
        heading_row_num = 1
        for row_itr in range(1,sheet_obj.max_row+1):
            cell_itr = sheet_obj.cell(row=row_itr,column=1).value
            if tc_name == cell_itr:
                for col_itr in range(1,sheet_obj.max_column+1):
                    key = sheet_obj.cell(row=heading_row_num,column=col_itr).value
                    if key == "Pre-conditions":
                        continue
                    if sheet_obj.cell(row=row_itr,column=col_itr).value == None:
                        excel_dict[key] = sheet_obj.cell(row=row_itr,column=col_itr).value
                    else:
                        excel_dict[key] = str(sheet_obj.cell(row=row_itr,column=col_itr).value)
                match_flag = 1
        return excel_dict
    except:
        import sys
        traceback.print_exception(*sys.exc_info())   
        return {}
