"""
Test case generator service for creating test cases from requirements
"""
import pandas as pd
import os
import sys
import json
from typing import List, Dict, Any, Optional
import logging
from langchain_core.prompts import PromptTemplate

# Adjust import path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from models.test_case import TestCase

logger = logging.getLogger(__name__)

class TestCaseGenerator:
    """Service for generating test cases from requirements"""
    
    def __init__(self, llm, retriever):
        """Initialize the generator with LLM and retriever"""
        self.llm = llm
        self.retriever = retriever
        self.logger = logging.getLogger(__name__)
        self._create_prompt_templates()
    
    def _create_prompt_templates(self):
        """Create prompt templates for test case generation"""
        self.extraction_prompt = PromptTemplate(
            template="""You are an expert test case writer specialized in converting software requirements into comprehensive test cases.

SIMILAR TEST CASES FOR REFERENCE:
{context}

REQUIREMENT TO ANALYZE:
{requirement}

Extract key requirements that need to be tested from the requirement document. Focus on functionalities, business rules, validations, and edge cases.

Your response should be a JSON list with each item containing:
- "requirement_id": An identifier for this requirement (derive from context if possible)
- "description": Brief description of the requirement 
- "testable_aspects": List of specific aspects that should be tested
- "criticality": HIGH, MEDIUM, or LOW

Format as valid JSON.
""",
            input_variables=["context", "requirement"]
        )
        
        self.generation_prompt = PromptTemplate(
            template="""You are an expert test case writer specialized in converting software requirements into comprehensive test cases.

SIMILAR TEST CASES FOR REFERENCE:
{context}

REQUIREMENT TO ANALYZE:
{requirement}

EXTRACTED TESTABLE ASPECTS:
{extracted_aspects}

Based on the extracted testable aspects, create complete test cases. Each test case should include:
1. A clear test case ID (format: TC_REQ_001, TC_REQ_002, etc.)
2. Test case description
3. Preconditions/setup requirements
4. Step-by-step test procedure
5. Expected results for each step
6. Test data needed
7. Priority (High/Medium/Low)

Follow the format and structure of the reference test cases where applicable.
Generate complete, executable test cases that could be followed by a tester.

Format your response as a JSON array with test case objects. Each test case should have these fields:
- test_case_id: string
- description: string
- preconditions: string
- steps: array of objects with "step_num" and "step_text"
- expected_results: array of strings (matching step numbers)
- test_data: string
- priority: string (High/Medium/Low)

Format as valid JSON.
""",
            input_variables=["context", "requirement", "extracted_aspects"]
        )
    
    def generate_test_cases(self, requirement_text: str) -> List[TestCase]:
        """Generate test cases from requirement text"""
        try:
            # Get similar test cases from vector store if retriever is available
            similar_test_cases = ""
            if self.retriever:
                similar_docs = self.retriever.get_relevant_documents(requirement_text)
                similar_test_cases = "\n\n".join([doc.page_content for doc in similar_docs])
            
            return self._generate_test_cases_internal(requirement_text, similar_test_cases)
                
        except Exception as e:
            self.logger.error(f"Error generating test cases: {str(e)}")
            return []
    
    def generate_test_cases_with_context(self, requirement_text: str, specific_context: str) -> List[TestCase]:
        """Generate test cases with specific context provided"""
        try:
            return self._generate_test_cases_internal(requirement_text, specific_context)
        except Exception as e:
            self.logger.error(f"Error generating test cases with context: {str(e)}")
            return []
    
    def _generate_test_cases_internal(self, requirement_text: str, context: str) -> List[TestCase]:
        """Internal method to generate test cases with given context"""
        try:
            # Extract testable aspects from requirement
            extraction_chain = self.llm.invoke(
                self.extraction_prompt.format(
                    context=context,
                    requirement=requirement_text
                )
            )
            # Extract content from AIMessage if it's an AIMessage object
            if hasattr(extraction_chain, 'content'):
                extracted_aspects = extraction_chain.content
            else:
                extracted_aspects = str(extraction_chain)
            
            # Generate test cases
            generation_chain = self.llm.invoke(
                self.generation_prompt.format(
                    context=context,
                    requirement=requirement_text,
                    extracted_aspects=extracted_aspects
                )
            )
            
            # Extract content from AIMessage if it's an AIMessage object
            if hasattr(generation_chain, 'content'):
                generation_response = generation_chain.content
            else:
                generation_response = str(generation_chain)
            
            # Parse the generated test cases
            try:
                # Clean the response by removing markdown code blocks if present
                cleaned_response = generation_response.strip()
                if cleaned_response.startswith('```json'):
                    # Remove markdown code block formatting
                    cleaned_response = cleaned_response[7:]  # Remove ```json
                    if cleaned_response.endswith('```'):
                        cleaned_response = cleaned_response[:-3]  # Remove trailing ```
                elif cleaned_response.startswith('```'):
                    # Remove generic code block formatting
                    cleaned_response = cleaned_response[3:]  # Remove ```
                    if cleaned_response.endswith('```'):
                        cleaned_response = cleaned_response[:-3]  # Remove trailing ```
                
                cleaned_response = cleaned_response.strip()
                
                generated_data = json.loads(cleaned_response)
                test_cases = []
                
                for tc_data in generated_data:
                    test_case = TestCase(
                        test_case_id=tc_data.get("test_case_id", ""),
                        description=tc_data.get("description", ""),
                        steps=tc_data.get("steps", []),
                        expected_results=tc_data.get("expected_results", []),
                        priority=tc_data.get("priority", "Medium"),
                        complexity=tc_data.get("complexity", "Medium"),
                        regression=tc_data.get("regression", "N"),
                        automated=tc_data.get("automated", "N"),
                        test_type=tc_data.get("test_type", "Manual"),
                        condition=tc_data.get("preconditions", ""),  # Map preconditions to condition
                        author=tc_data.get("author", "RAG Generated"),
                    )
                    test_cases.append(test_case)
                    
                return test_cases
            except json.JSONDecodeError as e:
                self.logger.error(f"Error parsing generated test cases: {str(e)}")
                self.logger.error(f"Raw response: {generation_response}")
                self.logger.error(f"Cleaned response: {cleaned_response if 'cleaned_response' in locals() else 'Not cleaned'}")
                return []
                
        except Exception as e:
            self.logger.error(f"Error in internal test case generation: {str(e)}")
            return []
    
    def export_to_excel(self, test_cases: List[TestCase], output_path: str) -> bool:
        """Export test cases to Excel format with merged cells"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create a new workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Cases"
            
            # Define headers
            headers = [
                "Test Case ID", "Test Case Description", "Step #", 
                "Test Step", "Test Case Expected Results", "Priority", 
                "Complexity", "Regression", "Automated", "Type", 
                "Author", "Condition"
            ]
            
            # Add headers to the worksheet
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Current row position
            current_row = 2
            
            # Process each test case
            for tc in test_cases:
                # Calculate how many rows this test case will need
                num_steps = len(tc.steps)
                start_row = current_row
                
                # Add data for each step
                for i, step in enumerate(tc.steps):
                    # Column A: Test Case ID (only for first row)
                    if i == 0:
                        cell_a = ws.cell(row=current_row, column=1, value=tc.test_case_id)
                        cell_a.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
                    
                    # Column B: Test Case Description (will be merged later)
                    if i == 0:
                        ws.cell(row=current_row, column=2, value=tc.description)
                    
                    # Column C: Step #
                    ws.cell(row=current_row, column=3, value=step.get("step_num", i+1))
                    
                    # Column D: Test Step
                    ws.cell(row=current_row, column=4, value=step.get("step_text", ""))
                    
                    # Column E: Expected Results
                    expected_result = tc.expected_results[i] if i < len(tc.expected_results) else ""
                    ws.cell(row=current_row, column=5, value=expected_result)
                    
                    # Other columns (only for first row)
                    if i == 0:
                        ws.cell(row=current_row, column=6, value=tc.priority)
                        ws.cell(row=current_row, column=7, value=tc.complexity)
                        ws.cell(row=current_row, column=8, value=tc.regression)
                        ws.cell(row=current_row, column=9, value=tc.automated)
                        ws.cell(row=current_row, column=10, value=tc.test_type)
                        ws.cell(row=current_row, column=11, value=tc.author)
                        ws.cell(row=current_row, column=12, value=tc.condition)
                    
                    current_row += 1
                
                # Merge cells for Test Case ID if multiple steps
                if num_steps > 1:
                    # Merge Test Case ID cells
                    ws.merge_cells(f'A{start_row}:A{start_row + num_steps - 1}')
                    # Merge Test Case Description cells
                    ws.merge_cells(f'B{start_row}:B{start_row + num_steps - 1}')
                    # Merge other metadata columns
                    for col in range(6, 13):  # Priority through Condition
                        col_letter = get_column_letter(col)
                        ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{start_row + num_steps - 1}')
            
            # Apply alignment to merged cells
            for row in ws.iter_rows(min_row=2, max_row=current_row-1):
                for cell in row:
                    if cell.column in [1, 2] or cell.column >= 6:  # ID, Description, and metadata columns
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    else:  # Step columns
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Apply yellow background to first column (Test Case ID)
            for row in range(2, current_row):
                cell = ws.cell(row=row, column=1)
                if cell.value:  # Only apply to cells that have content
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Auto-adjust column widths
            column_widths = {
                1: 15,  # Test Case ID
                2: 40,  # Test Case Description
                3: 8,   # Step #
                4: 50,  # Test Step
                5: 40,  # Expected Results
                6: 12,  # Priority
                7: 12,  # Complexity
                8: 12,  # Regression
                9: 12,  # Automated
                10: 12, # Type
                11: 15, # Author
                12: 25  # Condition
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Save the workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            self.logger.error(f"Error exporting test cases to Excel: {str(e)}")
            # Fallback to pandas if openpyxl fails
            try:
                data = []
                for tc in test_cases:
                    for i, step in enumerate(tc.steps):
                        row = {
                            "Test Case ID": tc.test_case_id if i == 0 else "",
                            "Test Case Description": tc.description if i == 0 else "",
                            "Step #": step.get("step_num", i+1),
                            "Test Step": step.get("step_text", ""),
                            "Test Case Expected Results": tc.expected_results[i] if i < len(tc.expected_results) else "",
                            "Priority": tc.priority if i == 0 else "",
                            "Complexity": tc.complexity if i == 0 else "",
                            "Regression": tc.regression if i == 0 else "",
                            "Automated": tc.automated if i == 0 else "",
                            "Type": tc.test_type if i == 0 else "",
                            "Author": tc.author if i == 0 else "",
                            "Condition": tc.condition if i == 0 else "",
                        }
                        data.append(row)
                
                df = pd.DataFrame(data)
                df.to_excel(output_path, index=False)
                return True
            except Exception as fallback_error:
                self.logger.error(f"Fallback export also failed: {str(fallback_error)}")
                return False
